"""Deeper inspection of source files."""
import sys, csv
from pathlib import Path
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl, docx

LEG = Path(r"C:\2026\369 Alliance\1. Legislation")
WORKING = Path(r"C:\2026\369 Alliance\1. Working folder")

# 1. Full AS in use docx
print("="*80)
print("AS in use.docx — FULL")
print("="*80)
d = docx.Document(LEG / "AS in use.docx")
t = d.tables[0]
current_cat = ""
for r in t.rows:
    a = r.cells[0].text.strip()
    b = r.cells[1].text.strip()
    if a == b:
        current_cat = a
        print(f"\n[{current_cat}]")
    else:
        print(f"  AS {a}  ::  {b}")

# 2. Reference Guide - sheet names
print("\n" + "="*80)
print("NCC Reference Guide — All sheets summary")
print("="*80)
wb = openpyxl.load_workbook(LEG / "NCC Building Reference Guide-Updated with new links 21 October 2021.xlsx", data_only=True)
for sn in wb.sheetnames:
    ws = wb[sn]
    print(f"  {sn}: {ws.max_row}r x {ws.max_column}c")

# 3. Examine 2019 sheet headers fully
print("\n--- 2019 sheet headers (row 1) ---")
ws = wb["2019"]
for col_idx in range(1, ws.max_column + 1):
    v = ws.cell(row=1, column=col_idx).value
    if v: print(f"  Col {col_idx}: {str(v)[:80]}")

# 4. Sample category rows
print("\n--- 2019 sheet sample categories (col A first 200 rows) ---")
for i in range(1, min(200, ws.max_row)+1):
    v = ws.cell(row=i, column=1).value
    if v and isinstance(v, str) and v.startswith("*"):
        print(f"  R{i}: {v[:100]}")
wb.close()

# 5. Look at NCC 2022 all clause file
print("\n" + "="*80)
print("NCC_2022_HP_allClause — full sample")
print("="*80)
with open(LEG / "NCC" / "NCC_2022_HP_allClause.csv", 'r', encoding='utf-8') as f:
    reader = csv.reader(f)
    rows = list(reader)
print(f"Total rows: {len(rows)}")
for i, r in enumerate(rows[:30]):
    print(f"  R{i}: {r[0][:120] if r else ''}")
print("...")
for i in range(len(rows)-10, len(rows)):
    print(f"  R{i}: {rows[i][0][:120] if rows[i] else ''}")

# 6. NCC Volume 2 Standards
print("\n" + "="*80)
print("NCC Volume 2 Standards — full")
print("="*80)
with open(LEG / "NCC" / "NCC Volume 2 Standards.csv", 'r', encoding='utf-8') as f:
    rows = list(csv.reader(f))
print(f"Total: {len(rows)} rows")
for r in rows[:50]:
    if r: print(f"  {r[0][:120]}")

# 7. Defect library full headers and stats
print("\n" + "="*80)
print("Defect Library — headers & categories")
print("="*80)
wb = openpyxl.load_workbook(WORKING / "New Master_Defect Library iAuditor_r.xlsx", data_only=True)
ws = wb.active
print(f"Sheet: {ws.title}, {ws.max_row}r x {ws.max_column}c")
for col_idx in range(1, ws.max_column + 1):
    v = ws.cell(row=1, column=col_idx).value
    print(f"  Col {col_idx}: {v}")
# unique categories
cats = set()
subcats = set()
for r in range(2, ws.max_row+1):
    cats.add(ws.cell(r, 2).value)
    subcats.add(ws.cell(r, 3).value)
print(f"\nUnique building elements: {sorted([c for c in cats if c])}")
print(f"\nUnique sub-categories ({len(subcats)}): {sorted([c for c in subcats if c])[:30]}")
wb.close()

# 8. Section CSVs available for NCC 2022
print("\n--- NCC 2022 Clauses CSV files ---")
for f in sorted((LEG/"NCC").glob("NCC_2022_Clauses_*.csv"))[:10]:
    print(f"  {f.name}")
print(f"  ... total: {len(list((LEG/'NCC').glob('NCC_2022_Clauses_*.csv')))}")
