"""
369 Alliance — Master NCC + AS Legislation Reference (v2)
=========================================================

v2 ADDS:
  • Full clause text from the NCC 2022 Volume One PDF (824 clauses)
  • Legacy 2019 reference column extracted from PDF "2019 X.Y" markers
  • PDF page-number column for easy cross-reference
  • Performance-Requirements-only filtered view
  • Deemed-to-Satisfy-only filtered view
  • Verification-Methods-only filtered view
  • Historical clause datasets — NCC 2015, NCC 2016, NCC 2019
  • Cross-edition longitudinal mapping (2022 ↔ 2019 ↔ 2016 ↔ 2015)

Sources:
  • All v1 sources, plus
  • scripts/cache/ncc2022_clause_texts.json (built by extract_ncc_pdf_text.py)
  • NCC2015_Section[A-J].csv          (NCC 2015 part-level)
  • NCC2016A1 SECTION [A-J].csv       (NCC 2016 Amendment 1)
  • NCC_2019A1_*.csv                  (NCC 2019 Amendment 1, 50 part files)
"""
from __future__ import annotations
import csv
import json
import re
import sys
from pathlib import Path
from datetime import datetime

sys.stdout.reconfigure(encoding="utf-8")

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import docx

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
ROOT       = Path(r"C:\2026\369 Alliance")
LEG        = ROOT / "1. Legislation"
NCC        = LEG / "NCC"
ONEDRIVE   = Path(r"C:\Users\raque\OneDrive\1. Working folder")
WORK_LEGACY = ROOT / "1. Working folder"
# Prefer OneDrive working folder; fall back to legacy on-disk location.
WORK = ONEDRIVE if ONEDRIVE.exists() else WORK_LEGACY
CACHE = Path(__file__).parent / "cache"
OUT   = WORK / "MASTER_NCC_AS_Legislation_Reference.xlsx"

CLAUSE_TEXTS = json.loads((CACHE / "ncc2022_clause_texts.json").read_text(encoding="utf-8"))

def find_first(*candidates: Path) -> Path | None:
    for p in candidates:
        if p.exists():
            return p
    return None

DEFECT_LIB_CANDIDATES = [
    ONEDRIVE / "New Master_Defect Library iAuditor_r.xlsx",
    WORK_LEGACY / "New Master_Defect Library iAuditor_r.xlsx",
    ROOT / "Library" / "2025 New Library" / "New Master_Defect Library iAuditor_r.xlsx",
    ROOT / "Bridge" / "DBP_Context_Bridge" / "2025 New Library" / "New Master_Defect Library iAuditor_r.xlsx",
]
DEFECTS_GENERAL_CANDIDATES = [
    ONEDRIVE / "Defects general.xlsx",
    WORK_LEGACY / "Defects general.xlsx",
    ROOT / "Library" / "2025 New Library" / "Defects general.xlsx",
]

# Sanitize PDF text for Excel: drop control chars, drop PUA glyphs, normalise
ILLEGAL_RX = re.compile(
    r"[\x00-\x08\x0B-\x0C\x0E-\x1F"   # ASCII control chars
    r"-"                   # private-use area (PDF ligature glyphs)
    r"￾￿]"
)
def sanitize(text: str) -> str:
    if not text: return ""
    t = ILLEGAL_RX.sub("", text)
    t = t.replace("�", "")  # replacement char from bad font extraction
    # Strip any "DRAFT" watermark fragments left over
    t = re.sub(r"(?m)^\s*DRAFT\s*$", "", t)
    t = re.sub(r"\n{3,}", "\n\n", t).strip()
    return t

# Apply sanitisation to all clause texts up-front
for _k, _v in CLAUSE_TEXTS.items():
    _v["text"] = sanitize(_v.get("text", ""))

# ---------------------------------------------------------------------------
# Brand styling (369 Alliance palette)
# ---------------------------------------------------------------------------
NAVY      = "1A1A2E"
GOLD      = "A68A64"
AMBER     = "C07040"
CREAM     = "F8F4ED"
LIGHTGOLD = "EFE3CC"
WHITE     = "FFFFFF"
GREY_HDR  = "2C2C40"
GREY_LITE = "F2F2F2"

SECTION_COLOURS = {
    "A": "5B6C8C", "B": "8B5A3C", "C": "B22222", "D": "FF8C00",
    "E": "DAA520", "F": "1E90FF", "G": "228B22", "H": "9370DB",
    "I": "9370DB", "J": "20B2AA", "S": "696969", "V2": "4F6228",
}
REGIME_COLOURS = {
    "Waterproofing": "1E90FF", "Fire safety": "B22222", "Fire": "B22222",
    "Structure": "8B5A3C",
    "Internal or external load-bearing component": "8B5A3C",
    "Building enclosure": "DAA520",
    "Mechanical, plumbing and electrical service": "20B2AA",
    "Essential Services": "20B2AA", "Cladding": "C07040",
}
CLAUSE_TYPE_COLOURS = {
    "Objective":            "B0C4DE",
    "Functional Statement": "C2E0FF",
    "Performance Requirement": "F4B084",
    "Verification Method":  "FFD966",
    "Deemed-to-Satisfy":    "C6EFCE",
    "Specification Clause": "D9D9D9",
    "Governing / General":  "E7E6E6",
}

THIN = Side(style="thin", color="BFBFBF")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER  = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT    = Alignment(horizontal="left",   vertical="center", wrap_text=True)
LEFTTOP = Alignment(horizontal="left",   vertical="top",    wrap_text=True)

def hdr_fill(hex_): return PatternFill("solid", fgColor=hex_)
def font(hex_=None, bold=False, size=11, italic=False):
    kw = dict(bold=bold, size=size, italic=italic, name="Calibri")
    if hex_: kw["color"] = hex_
    return Font(**kw)

def write_header_row(ws, headers, fill_hex=NAVY, font_hex=WHITE, row_idx=1):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row_idx, column=c, value=h)
        cell.fill = hdr_fill(fill_hex)
        cell.font = font(font_hex, bold=True, size=11)
        cell.alignment = CENTER
        cell.border = BORDER_ALL
    ws.row_dimensions[row_idx].height = 36

def autosize(ws, max_widths=None):
    max_widths = max_widths or {}
    for col_idx, col in enumerate(ws.columns, 1):
        letter = get_column_letter(col_idx)
        max_len = 10
        for cell in col:
            try:
                v = str(cell.value) if cell.value is not None else ""
            except Exception:
                v = ""
            for line in v.split("\n"):
                if len(line) > max_len:
                    max_len = len(line)
        cap = max_widths.get(letter, 60)
        ws.column_dimensions[letter].width = min(max_len + 2, cap)

def add_filter_and_freeze(ws, header_row=1, freeze_at=None):
    last_col = get_column_letter(ws.max_column)
    ws.auto_filter.ref = f"A{header_row}:{last_col}{ws.max_row}"
    ws.freeze_panes = freeze_at or f"A{header_row+1}"

# ---------------------------------------------------------------------------
# Source loaders
# ---------------------------------------------------------------------------
def load_csv_labels(path: Path) -> list[str]:
    out = []
    with open(path, encoding="utf-8") as f:
        for i, row in enumerate(csv.reader(f)):
            if i == 0 and row and row[0].lstrip("﻿").strip().lower() == "label":
                continue
            if not row:
                continue
            text = " ".join([c for c in row if c]).strip().lstrip("﻿").strip()
            if text:
                out.append(text)
    return out

# ---- NCC 2022 Vol 1 ----
CLAUSE_VOL1 = re.compile(r"^(?P<code>[A-Z]+\d*[A-Z]?\d*)\s+(?P<title>.+)$")
CLAUSE_TYPE_MAP = {
    "O": "Objective", "F": "Functional Statement", "P": "Performance Requirement",
    "V": "Verification Method", "D": "Deemed-to-Satisfy",
    "G": "Governing / General", "C": "Specification Clause",
}

def parse_vol1_filename(fname: str) -> tuple[str, str, str, str]:
    base = fname.replace(".csv", "")
    m = re.search(r"Specification_(\d+)_(.+)$", base)
    if m:
        return ("S", f"S{m.group(1)}", f"Specification {m.group(1)} — {m.group(2).replace('_', ' ')}", "Specification")
    m = re.search(r"Part_([A-Z]\d?)_?(.*)$", base)
    if m:
        title = m.group(2).replace("_", " ").strip() or m.group(1)
        return (m.group(1)[0], m.group(1), f"Part {m.group(1)} — {title}", "Part")
    return ("?", "?", base, "?")

def classify_clause_code(code: str) -> str:
    m = re.match(r"^[A-Z]+\d+([A-Z])\d+", code)
    if m: return CLAUSE_TYPE_MAP.get(m.group(1), "Provision")
    m = re.match(r"^S\d+([A-Z])\d+", code)
    if m: return CLAUSE_TYPE_MAP.get(m.group(1), "Specification Clause")
    return "Provision"

LEGACY_REF_RX = re.compile(r"^\s*(?:20\d{2})\s+([A-Z]+\.?[\w\.\-]*[\w])", re.MULTILINE)

def extract_legacy_ref(clause_text: str) -> str:
    if not clause_text: return ""
    m = LEGACY_REF_RX.search(clause_text[:600])
    return m.group(1).strip() if m else ""

def build_ncc_vol1():
    rows = []
    for f in sorted(NCC.glob("NCC_2022_Clauses_*.csv")):
        section, part_code, part_title, kind = parse_vol1_filename(f.name)
        for raw in load_csv_labels(f):
            m = CLAUSE_VOL1.match(raw)
            code = m.group("code") if m else ""
            title = m.group("title") if m else raw
            ctype = classify_clause_code(code) if code else ""
            tx = CLAUSE_TEXTS.get(code, {})
            rows.append({
                "ncc_version": "NCC 2022", "volume": "Volume One",
                "section": section, "part_code": part_code, "part_title": part_title,
                "kind": kind,
                "clause_code": code, "clause_type": ctype, "clause_title": title,
                "full_text":   tx.get("text", ""),
                "pdf_page":    tx.get("page", ""),
                "legacy_2019": extract_legacy_ref(tx.get("text", "")),
                "source_file": f.name,
            })
    return rows

# ---- NCC 2022 Vol 2 (HP) ----
HP_PARSE = re.compile(
    r"^(?P<sec>\d+)\s+(?P<sec_title>[^-–>]+?)\s*[-–>]+\s*"
    r"(?:Part\s+(?P<part>\d+(?:\.\d+)?)\s*(?P<part_title>[^-–>]*?))?"
    r"\s*(?:[-–>]+\s*(?P<clause>\d+(?:\.\d+)+)\s*(?P<clause_title>.+))?$"
)

def build_ncc_vol2():
    rows = []
    for raw in load_csv_labels(NCC / "NCC_2022_HP_allClause.csv"):
        clean = raw.replace(" > ", " - ").strip()
        m = HP_PARSE.match(clean)
        if m:
            rows.append({
                "ncc_version": "NCC 2022", "volume": "Volume Two (Housing Provisions)",
                "section": m.group("sec") or "", "section_title": (m.group("sec_title") or "").strip(),
                "part_code": (m.group("part") or "").strip(), "part_title": (m.group("part_title") or "").strip(),
                "clause_code": (m.group("clause") or "").strip(),
                "clause_title": (m.group("clause_title") or m.group("part_title") or m.group("sec_title") or "").strip(),
                "source": "NCC_2022_HP_allClause.csv",
            })
        else:
            rows.append({
                "ncc_version": "NCC 2022", "volume": "Volume Two (Housing Provisions)",
                "section": "", "section_title": "", "part_code": "", "part_title": "",
                "clause_code": "", "clause_title": clean, "source": "NCC_2022_HP_allClause.csv",
            })
    return rows

# ---- Historical NCC 2019 Vol 1 ----
NCC_2019_PARTS = list(NCC.glob("NCC_2019A1_*.csv"))

def parse_2019_filename(name: str) -> tuple[str, str, str, str]:
    base = name.replace(".csv", "")
    m = re.search(r"Part\s+([A-Z]\d?)\s+(.+?)(?:\s*\((DtS)\))?\s*$", base)
    if m:
        part = m.group(1)
        title = m.group(2).strip()
        section = part[0]
        kind = "Part (DtS)" if (m.group(3) or "").lower() == "dts" else "Part"
        return (section, part, title, kind)
    m = re.search(r"Section\s+([A-Z])\s+(.+)$", base)
    if m:
        return (m.group(1), m.group(1), m.group(2), "Section")
    return ("?", "?", base, "?")

def build_ncc_2019():
    rows = []
    for f in sorted(NCC_2019_PARTS):
        section, part, part_title, kind = parse_2019_filename(f.name)
        for raw in load_csv_labels(f):
            m = re.match(r"^([A-Z]+\d*\.?\d*)\s+(.+)$", raw)
            code  = m.group(1) if m else ""
            title = m.group(2) if m else raw
            rows.append({
                "ncc_version": "NCC 2019",
                "section": section, "part": part, "part_title": part_title, "kind": kind,
                "clause_code": code, "clause_title": title,
                "source": f.name,
            })
    return rows

# ---- Historical NCC 2016 Vol 1 ----
def build_ncc_2016():
    rows = []
    for f in sorted(NCC.glob("NCC2016A1 SECTION *.csv")):
        section_letter = f.name.split("SECTION ")[1].split(".")[0].strip()
        part_now = ""
        for raw in load_csv_labels(f):
            if raw.upper().startswith("PART "):
                part_now = raw.replace("PART ", "Part ").strip()
                continue
            m = re.match(r"^([A-Z]+\d*\.?\d*)\s+(.+)$", raw)
            code  = m.group(1) if m else ""
            title = m.group(2) if m else raw
            rows.append({
                "ncc_version": "NCC 2016 (A1)",
                "section": section_letter, "part": part_now,
                "clause_code": code, "clause_title": title,
                "source": f.name,
            })
    return rows

# ---- Historical NCC 2015 Vol 1 ----
def build_ncc_2015():
    rows = []
    for f in sorted(NCC.glob("NCC2015_Section*.csv")):
        section_letter = f.stem.replace("NCC2015_Section", "").strip()
        if section_letter.lower() == "s": continue
        part_now = ""
        for raw in load_csv_labels(f):
            if raw.upper().startswith("PART "):
                part_now = raw
                continue
            m = re.match(r"^([A-Z]+\d*\.?\d*)\s+(.+)$", raw)
            code  = m.group(1) if m else ""
            title = m.group(2) if m else raw
            rows.append({
                "ncc_version": "NCC 2015",
                "section": section_letter, "part": part_now,
                "clause_code": code, "clause_title": title,
                "source": f.name,
            })
    return rows

# ---- Standards & rest (re-used from v1) ----
STD_PARSE = re.compile(r"^\[(?P<vstr>NCC\d{4})\]\s*(?P<rest>.+)$")

def build_referenced_standards():
    rows = []
    with open(NCC / "NCC Volume 2 Standards.csv", encoding="utf-8") as f:
        for i, row in enumerate(csv.reader(f)):
            if i == 0 or not row: continue
            text = " ".join([c.strip() for c in row if c]).lstrip("﻿").strip()
            if not text: continue
            m = STD_PARSE.match(text)
            if not m:
                rows.append({"ncc_version": "Unknown", "code": "", "year": "",
                             "title": text, "version_note": "", "source": "NCC Volume 2 Standards.csv"})
                continue
            parts = [p.strip() for p in m.group("rest").split("|")]
            rows.append({
                "ncc_version": m.group("vstr"),
                "code":  parts[0] if parts else "",
                "year":  parts[1] if len(parts) > 1 else "",
                "title": parts[2] if len(parts) > 2 else "",
                "version_note": parts[3] if len(parts) > 3 else "",
                "source": "NCC Volume 2 Standards.csv",
            })
    return rows

def build_as_in_use():
    rows = []
    p = find_first(
        LEG / "AS in use.docx",
        ONEDRIVE / "AS in use.docx",
        WORK_LEGACY / "AS in use.docx",
    )
    if not p:
        print("[!] AS in use.docx not found")
        return rows
    d = docx.Document(p)
    cat = ""
    for r in d.tables[0].rows:
        a, b = r.cells[0].text.strip(), r.cells[1].text.strip()
        if a == b and a: cat = a
        elif a:          rows.append({"category": cat, "code": a, "title": b})
    return rows

def build_crosswalk():
    rows = []
    p = find_first(
        LEG / "NCC Building Reference Guide-Updated with new links 21 October 2021.xlsx",
        ONEDRIVE / "NCC Building Reference Guide-Updated with new links 21 October 2021.xlsx",
        WORK_LEGACY / "NCC Building Reference Guide-Updated with new links 21 October 2021.xlsx",
    )
    if not p:
        print("[!] NCC Building Reference Guide not found")
        return rows
    wb = openpyxl.load_workbook(p, data_only=True)
    if "2019" not in wb.sheetnames:
        wb.close(); return rows
    ws = wb["2019"]
    topic = ""
    for ridx in range(2, ws.max_row + 1):
        item = ws.cell(ridx, 1).value
        if item is None: continue
        if isinstance(item, str) and item.strip().startswith("*"):
            topic = item.strip().lstrip("*").strip(); continue
        if not isinstance(item, str) or not item.strip(): continue
        rows.append({
            "topic": topic, "complaint_item": item.strip(),
            "ncc_vol1_2019":  (ws.cell(ridx, 2).value or "").strip() if ws.cell(ridx, 2).value else "",
            "ncc_vol2_2019":  (ws.cell(ridx, 3).value or "").strip() if ws.cell(ridx, 3).value else "",
            "ncc_vol3_2019":  (ws.cell(ridx, 4).value or "").strip() if ws.cell(ridx, 4).value else "",
            "as_reference":   (ws.cell(ridx, 5).value or "").strip() if ws.cell(ridx, 5).value else "",
            "guide_to_std":   (ws.cell(ridx, 6).value or "").strip() if ws.cell(ridx, 6).value else "",
            "other":          (ws.cell(ridx, 7).value or "").strip() if ws.cell(ridx, 7).value else "",
            "previous":       (ws.cell(ridx, 8).value or "").strip() if ws.cell(ridx, 8).value else "",
        })
    wb.close()
    return rows

def build_defects():
    rows = []
    p = find_first(*DEFECT_LIB_CANDIDATES)
    if not p:
        print("[!] Defect library file not found — skipping iAuditor defects")
        return rows
    wb = openpyxl.load_workbook(p, data_only=True)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    for r in range(2, ws.max_row + 1):
        rec = {h: ws.cell(r, c).value for c, h in enumerate(headers, 1)}
        rec = {k: v.strip() if isinstance(v, str) else v for k, v in rec.items()}
        if rec.get("Defect ID"): rows.append(rec)
    wb.close()
    return rows

# ---- Defects general.xlsx (NSW Building Commissioner element codes + 2019/2022 register) ----
SHORT_CODES_FALLBACK = [
    ("W1",  "Waterproofing",      "External waterproofing"),
    ("W2",  "Waterproofing",      "Internal waterproofing"),
    ("W3",  "Waterproofing",      "Other"),
    ("W4",  "Waterproofing",      "Other / not in library"),
    ("F1",  "Fire Safety",        "Fire resistance and stability"),
    ("F2",  "Fire Safety",        "Compartmentation"),
    ("F3",  "Fire Safety",        "Protection of openings"),
    ("F4",  "Fire Safety",        "Access and Egress"),
    ("F5",  "Fire Safety",        "Fire fighting equipment and wet systems"),
    ("F6",  "Fire Safety",        "Smoke Hazard management and detection"),
    ("F7",  "Fire Safety",        "Lift installation"),
    ("F8",  "Fire Safety",        "Exit signs, emergency lights and warning"),
    ("F9",  "Fire Safety",        "Other"),
    ("S1",  "Structure",          "Concrete"),
    ("S2",  "Structure",          "Steel"),
    ("S3",  "Structure",          "Masonry/Blockwork"),
    ("S4",  "Structure",          "Timber"),
    ("S5",  "Structure",          "Other"),
    ("BE1", "Building Enclosure", "Cavity Masonry"),
    ("BE2", "Building Enclosure", "Proprietary Cladding"),
    ("BE3", "Building Enclosure", "Windows, Doors and Glazing"),
    ("BE4", "Building Enclosure", "Other"),
    ("BE5", "Building Enclosure", "Other / Roofing"),
    ("ES1", "Essential Service",  "Sanitary drainage"),
    ("ES2", "Essential Service",  "Stormwater drainage"),
    ("ES3", "Essential Service",  "Electrical"),
    ("ES4", "Essential Service",  "Mechanical and ventilation"),
    ("ES5", "Essential Service",  "Gas plumbing"),
    ("ES6", "Essential Service",  "Vertical transportation"),
    ("ES7", "Essential Service",  "Other"),
]

def build_short_codes():
    rows = []
    p = find_first(*DEFECTS_GENERAL_CANDIDATES)
    if not p:
        return [{"code": c, "element": e, "description": d} for c, e, d in SHORT_CODES_FALLBACK]
    wb = openpyxl.load_workbook(p, data_only=True)
    ws = wb["ShortCodes"]
    for r in range(2, ws.max_row + 1):
        code = ws.cell(r, 2).value
        elem = ws.cell(r, 3).value
        desc = ws.cell(r, 4).value
        if code in (None, "Element Codes"): continue
        if not (code or elem): continue
        rows.append({
            "code":        (code or "").strip() if isinstance(code, str) else code,
            "element":     (elem or "").strip() if isinstance(elem, str) else "",
            "description": (desc or "").strip() if isinstance(desc, str) else "",
        })
    wb.close()
    return rows

def build_defects_general(sheet_name: str):
    rows = []
    p = find_first(*DEFECTS_GENERAL_CANDIDATES)
    if not p:
        return rows
    wb = openpyxl.load_workbook(p, data_only=True)
    ws = wb[sheet_name]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    for r in range(2, ws.max_row + 1):
        rec = {}
        for c, h in enumerate(headers, 1):
            v = ws.cell(r, c).value
            if isinstance(v, str): v = v.strip()
            rec[h] = v
        if rec.get("Unique ID") or rec.get("Element Code"):
            rows.append(rec)
    wb.close()
    return rows

# ---- Canonical NCC ↔ AS relationship register (NCC and AS.xlsx) ----
NCC_AS_CANDIDATES = [
    ONEDRIVE / "NCC and AS.xlsx",
    WORK_LEGACY / "NCC and AS.xlsx",
]

def build_ncc_as_register():
    """
    Headers live in row 2 (row 1 has the merged 'Legislation' group label).
    Returns list of dict rows.
    """
    p = find_first(*NCC_AS_CANDIDATES)
    if not p:
        print("[!] NCC and AS.xlsx not found — skipping NCC-AS register")
        return []
    rows = []
    wb = openpyxl.load_workbook(p, data_only=True)
    ws = wb["NCC 2022 (2)"]
    headers = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]
    for r in range(3, ws.max_row + 1):
        rec = {}
        for c, h in enumerate(headers, 1):
            if not h: continue
            v = ws.cell(r, c).value
            if isinstance(v, str): v = v.strip()
            rec[h] = v
        if rec.get("Unique ID") or rec.get("Element Code"):
            rows.append(rec)
    wb.close()
    return rows

# Parse a free-text NCC Quote field into individual clause tokens
NCC_CLAUSE_RX = re.compile(
    r"\b("
    r"S\d{1,3}C\d{1,3}"                       # S26C3
    r"|[A-Z]\d+[A-Z]\d+"                      # F1P2, B1D4, S26C3 (both styles)
    r"|Specification\s+\d+"                   # Specification 26
    r"|Part\s+[A-Z]\d?"                       # Part B1
    r")\b"
)

# Parse AS standard with optional year (loose match, then we normalise)
AS_STD_RX = re.compile(r"\bAS\s*(?:/\s*NZS)?\s*\d+(?:\.\d+)*(?:\s*[-:]\s*\d{4})?", re.IGNORECASE)

def normalise_as_code(s: str) -> str:
    """E.g. 'as 4654.2-2012' or '4654.2-2012' -> 'AS 4654.2-2012'."""
    if not s: return ""
    s = s.strip().rstrip(",;")
    if not s: return ""
    if s.upper().startswith("AS"):
        s = re.sub(r"^AS\s*", "AS ", s, flags=re.IGNORECASE)
        s = re.sub(r"\s+", " ", s)
        return s.upper().replace("AS/NZS", "AS/NZS")
    return f"AS {s}"

def build_ncc_as_relationship(register_rows: list[dict]) -> dict:
    """
    From the defect register, build:
      ncc_to_as[ncc_clause_code] = list of {as_std, as_clauses, defect_id, brief}
    """
    mapping = {}
    for rec in register_rows:
        ncc_quotes = str(rec.get("NCC Quotes") or "")
        as_used    = str(rec.get("AS Used") or "")
        as_clauses = str(rec.get("AS Clauses") or "")
        uid        = str(rec.get("Unique ID") or "")
        brief      = str(rec.get("Brief Description/What is the defect?") or "")
        if not ncc_quotes:
            continue

        # Extract NCC clause tokens
        ncc_codes = []
        for m in NCC_CLAUSE_RX.finditer(ncc_quotes):
            ncc_codes.append(m.group(1).strip())
        if not ncc_codes:
            continue

        # Extract AS standards (may be multiple)
        as_list = []
        if as_used and as_used != "-":
            # Split by comma, then normalise each
            for token in re.split(r"[,;]", as_used):
                token = token.strip()
                if not token or token == "-": continue
                as_list.append(normalise_as_code(token))
            if not as_list:
                as_list = [as_used.strip()]
        as_clauses_str = "" if (not as_clauses or as_clauses == "-") else as_clauses

        for code in ncc_codes:
            entry = mapping.setdefault(code, [])
            if as_list:
                for std in as_list:
                    entry.append({
                        "as": std,
                        "as_clauses": as_clauses_str,
                        "defect_id": uid,
                        "brief": brief,
                    })
            else:
                entry.append({
                    "as": "", "as_clauses": "",
                    "defect_id": uid, "brief": brief,
                })
    return mapping

def aggregate_as_for_clause(entries: list[dict]) -> tuple[str, str, str]:
    """Return (as_standards_csv, as_clauses_pretty, defect_ids_csv) for display."""
    if not entries:
        return ("", "", "")
    stds = sorted({e["as"] for e in entries if e["as"]})
    # Build "AS X-YYYY: clauses; AS Z-YYYY: clauses" pretty string
    by_std = {}
    for e in entries:
        if e["as"] and e["as_clauses"]:
            by_std.setdefault(e["as"], set()).update(
                p.strip() for p in re.split(r"[,;]", e["as_clauses"]) if p.strip()
            )
    pretty_parts = []
    for std, cls in by_std.items():
        if cls:
            pretty_parts.append(f"{std}: {', '.join(sorted(cls))}")
        else:
            pretty_parts.append(std)
    defect_ids = sorted({e["defect_id"] for e in entries if e["defect_id"]})
    return (
        ", ".join(stds),
        " | ".join(pretty_parts),
        ", ".join(defect_ids),
    )

# ---------------------------------------------------------------------------
# Sheet writers
# ---------------------------------------------------------------------------
def write_vol1_sheet(ws, vol1_rows, ncc_as_map):
    headers = ["NCC Version", "Volume", "Sec", "Part Code", "Part Title",
               "Kind", "NCC Clause", "Clause Type", "Clause Title",
               "Full Clause Text (NCC 2022 PDF)", "PDF Page", "Legacy 2019 Ref",
               "AS (Australian Standard)", "AS Clause", "Linked Defect IDs", "Source File"]
    write_header_row(ws, headers, NAVY, WHITE)
    for i, row in enumerate(vol1_rows, 2):
        ws.cell(i, 1, row["ncc_version"])
        ws.cell(i, 2, row["volume"])
        sc = ws.cell(i, 3, row["section"])
        sc.fill = hdr_fill(SECTION_COLOURS.get(row["section"], "BFBFBF"))
        sc.font = font(WHITE, bold=True); sc.alignment = CENTER
        ws.cell(i, 4, row["part_code"]).font = font(NAVY, bold=True)
        ws.cell(i, 5, row["part_title"])
        ws.cell(i, 6, row["kind"])
        code = row["clause_code"]
        ws.cell(i, 7, code).font = font(NAVY, bold=True)
        ct_cell = ws.cell(i, 8, row["clause_type"])
        c_col = CLAUSE_TYPE_COLOURS.get(row["clause_type"])
        if c_col:
            ct_cell.fill = hdr_fill(c_col); ct_cell.alignment = CENTER
        ws.cell(i, 9, row["clause_title"])
        text_cell = ws.cell(i, 10, row["full_text"])
        text_cell.alignment = LEFTTOP
        text_cell.font = font(GREY_HDR, size=9)
        ws.cell(i, 11, row["pdf_page"]).alignment = CENTER
        leg = ws.cell(i, 12, row["legacy_2019"])
        leg.font = font(AMBER, bold=True); leg.alignment = CENTER
        # NEW: AS / AS Clause / Defect IDs from canonical mapping
        as_std, as_cl_pretty, defect_ids = aggregate_as_for_clause(ncc_as_map.get(code, []))
        as_cell = ws.cell(i, 13, as_std)
        if as_std:
            as_cell.fill = hdr_fill("EFE3CC"); as_cell.font = font("8B5A3C", bold=True)
        as_cell.alignment = LEFTTOP
        ascl = ws.cell(i, 14, as_cl_pretty)
        if as_cl_pretty:
            ascl.font = font(GREY_HDR, italic=True, size=9)
        ascl.alignment = LEFTTOP
        defs = ws.cell(i, 15, defect_ids)
        if defect_ids:
            defs.font = font(AMBER, bold=True, size=9)
        defs.alignment = LEFTTOP
        ws.cell(i, 16, row["source_file"]).font = font(GREY_HDR, size=9, italic=True)
        for col in range(1, len(headers)+1):
            ws.cell(i, col).border = BORDER_ALL
            if col not in (3, 8, 11, 12):
                ws.cell(i, col).alignment = LEFTTOP
        ws.row_dimensions[i].height = 60 if row["full_text"] else 22
    autosize(ws, max_widths={"A": 12, "B": 14, "C": 6, "D": 10, "E": 38, "F": 14,
                             "G": 12, "H": 22, "I": 38, "J": 90, "K": 8, "L": 14,
                             "M": 30, "N": 50, "O": 24, "P": 30})
    add_filter_and_freeze(ws)

def write_filtered_view(ws, vol1_rows, predicate, header_fill, title_block, ncc_as_map):
    filtered = [r for r in vol1_rows if predicate(r)]
    headers = ["Sec", "Part Code", "NCC Clause", "Clause Title",
               "Full Clause Text", "PDF Page", "Legacy 2019 Ref",
               "AS (Australian Standard)", "AS Clause", "Linked Defect IDs"]
    write_header_row(ws, headers, header_fill, WHITE)
    for i, row in enumerate(filtered, 2):
        sc = ws.cell(i, 1, row["section"])
        sc.fill = hdr_fill(SECTION_COLOURS.get(row["section"], "BFBFBF"))
        sc.font = font(WHITE, bold=True); sc.alignment = CENTER
        ws.cell(i, 2, row["part_code"]).font = font(NAVY, bold=True)
        code = row["clause_code"]
        ws.cell(i, 3, code).font = font(NAVY, bold=True)
        ws.cell(i, 4, row["clause_title"])
        tc = ws.cell(i, 5, row["full_text"])
        tc.alignment = LEFTTOP; tc.font = font(GREY_HDR, size=9)
        ws.cell(i, 6, row["pdf_page"]).alignment = CENTER
        leg = ws.cell(i, 7, row["legacy_2019"])
        leg.font = font(AMBER, bold=True); leg.alignment = CENTER
        as_std, as_cl_pretty, defect_ids = aggregate_as_for_clause(ncc_as_map.get(code, []))
        ac = ws.cell(i, 8, as_std)
        if as_std: ac.fill = hdr_fill("EFE3CC"); ac.font = font("8B5A3C", bold=True)
        ac.alignment = LEFTTOP
        ws.cell(i, 9, as_cl_pretty).font = font(GREY_HDR, italic=True, size=9)
        ws.cell(i, 9).alignment = LEFTTOP
        ws.cell(i, 10, defect_ids).font = font(AMBER, bold=True, size=9)
        ws.cell(i, 10).alignment = LEFTTOP
        for col in range(1, len(headers)+1):
            ws.cell(i, col).border = BORDER_ALL
            if col not in (1, 6, 7):
                ws.cell(i, col).alignment = LEFTTOP
        ws.row_dimensions[i].height = 70 if row["full_text"] else 22
    autosize(ws, max_widths={"A": 6, "B": 10, "C": 14, "D": 40, "E": 100,
                             "F": 8, "G": 16, "H": 50, "I": 24, "J": 30})
    add_filter_and_freeze(ws)
    return len(filtered)

def write_historical_sheet(ws, rows, version_label, header_fill):
    headers = ["NCC Version", "Sec", "Part", "Part Title", "Kind",
               "Clause Code", "Clause Title", "Source File"]
    write_header_row(ws, headers, header_fill, WHITE)
    for i, row in enumerate(rows, 2):
        ws.cell(i, 1, row["ncc_version"]).alignment = CENTER
        sc = ws.cell(i, 2, row["section"])
        sc.fill = hdr_fill(SECTION_COLOURS.get(row["section"], "BFBFBF"))
        sc.font = font(WHITE, bold=True); sc.alignment = CENTER
        ws.cell(i, 3, row.get("part", "")).font = font(NAVY, bold=True)
        ws.cell(i, 4, row.get("part_title", ""))
        ws.cell(i, 5, row.get("kind", ""))
        ws.cell(i, 6, row.get("clause_code", "")).font = font(NAVY, bold=True)
        ws.cell(i, 7, row.get("clause_title", ""))
        ws.cell(i, 8, row.get("source", "")).font = font(GREY_HDR, size=9, italic=True)
        for col in range(1, len(headers)+1):
            ws.cell(i, col).border = BORDER_ALL
            if col not in (1, 2, 3, 6):
                ws.cell(i, col).alignment = LEFTTOP
    autosize(ws, max_widths={"A": 14, "B": 6, "C": 14, "D": 40, "E": 16, "F": 14, "G": 70, "H": 32})
    add_filter_and_freeze(ws)

def write_cross_edition(ws, vol1_rows, ncc2019_rows, ncc2016_rows, ncc2015_rows):
    """For each NCC 2022 Vol 1 clause, attempt to map to legacy editions
    using the 'Legacy 2019 Ref' extracted from PDF text + heuristic match."""
    # Build lookup tables for older editions: code -> title
    by2019 = {(r.get("clause_code") or "").strip(): r for r in ncc2019_rows if r.get("clause_code")}
    by2016 = {(r.get("clause_code") or "").strip(): r for r in ncc2016_rows if r.get("clause_code")}
    by2015 = {(r.get("clause_code") or "").strip(): r for r in ncc2015_rows if r.get("clause_code")}

    headers = ["NCC 2022 Code", "NCC 2022 Title", "Section",
               "Legacy 2019 Ref", "NCC 2019 Title", "NCC 2016 Title", "NCC 2015 Title"]
    write_header_row(ws, headers, "C07040", WHITE)
    rownum = 2
    for r in vol1_rows:
        code22 = r["clause_code"]
        if not code22: continue
        legacy = r.get("legacy_2019", "").strip()
        # Try direct match first, then strip trailing extras
        candidates = []
        if legacy:
            candidates.append(legacy)
            # Common variants: "BP1.1" vs "B1.1"; strip "P","V" letter
            stripped = re.sub(r"^([A-Z])[A-Z]", r"\1", legacy)
            if stripped != legacy: candidates.append(stripped)
        # Also try: 2022 code converted
        m = re.match(r"^([A-Z])(\d+)([A-Z])(\d+)$", code22)
        if m:
            sec, part, typ, num = m.groups()
            # 2019 P/V/D notation: BP1.1, BV1, B1.0
            if typ == "P":      candidates.append(f"{sec}P{part}.{num}")
            elif typ == "V":    candidates.append(f"{sec}V{num}")
            elif typ == "D":    candidates.append(f"{sec}{part}.{num}")
            elif typ == "O":    candidates.append(f"{sec}O{part}")
            elif typ == "F":    candidates.append(f"{sec}F{part}.{num}")

        t2019 = t2016 = t2015 = ""
        m2019_used = ""
        for c in candidates:
            if c in by2019 and not t2019:
                t2019 = by2019[c].get("clause_title", "")
                m2019_used = c
            if c in by2016 and not t2016:
                t2016 = by2016[c].get("clause_title", "")
            if c in by2015 and not t2015:
                t2015 = by2015[c].get("clause_title", "")
        # Skip rows where we found absolutely nothing useful
        if not (t2019 or t2016 or t2015 or legacy):
            continue
        ws.cell(rownum, 1, code22).font = font(NAVY, bold=True)
        ws.cell(rownum, 2, r["clause_title"])
        sc = ws.cell(rownum, 3, r["section"])
        sc.fill = hdr_fill(SECTION_COLOURS.get(r["section"], "BFBFBF"))
        sc.font = font(WHITE, bold=True); sc.alignment = CENTER
        ws.cell(rownum, 4, m2019_used or legacy).font = font(AMBER, bold=True)
        ws.cell(rownum, 5, t2019)
        ws.cell(rownum, 6, t2016)
        ws.cell(rownum, 7, t2015)
        for col in range(1, len(headers)+1):
            ws.cell(rownum, col).border = BORDER_ALL
            ws.cell(rownum, col).alignment = LEFTTOP
        rownum += 1
    autosize(ws, max_widths={"A": 16, "B": 50, "C": 6, "D": 16, "E": 50, "F": 50, "G": 50})
    add_filter_and_freeze(ws)
    return rownum - 2

# ---------------------------------------------------------------------------
# Main builder
# ---------------------------------------------------------------------------
def main():
    print("[ ] Loading sources …")
    vol1     = build_ncc_vol1();              print(f"    Vol 1 (2022) clauses          : {len(vol1)}")
    vol2     = build_ncc_vol2();              print(f"    Vol 2 (HP, 2022) clauses      : {len(vol2)}")
    refstd   = build_referenced_standards();  print(f"    Referenced standards          : {len(refstd)}")
    asuse    = build_as_in_use();             print(f"    AS in use (categorised)       : {len(asuse)}")
    cross    = build_crosswalk();             print(f"    Topic crosswalk rows          : {len(cross)}")
    defects  = build_defects();               print(f"    Defect library entries        : {len(defects)}")
    ncc2019  = build_ncc_2019();              print(f"    NCC 2019 clauses              : {len(ncc2019)}")
    ncc2016  = build_ncc_2016();              print(f"    NCC 2016 clauses              : {len(ncc2016)}")
    ncc2015  = build_ncc_2015();              print(f"    NCC 2015 clauses              : {len(ncc2015)}")
    short_codes  = build_short_codes();              print(f"    NSW BC element short codes    : {len(short_codes)}")
    nsw_def_2019 = build_defects_general("NCC 2019");print(f"    NSW defect register (2019)    : {len(nsw_def_2019)}")
    nsw_def_2022 = build_defects_general("NCC 2022");print(f"    NSW defect register (2022)    : {len(nsw_def_2022)}")
    ncc_as_reg   = build_ncc_as_register();          print(f"    NCC↔AS canonical register     : {len(ncc_as_reg)}")
    ncc_as_map   = build_ncc_as_relationship(ncc_as_reg)
    print(f"    NCC clauses with AS mapping   : {len(ncc_as_map)}")

    matched_text = sum(1 for r in vol1 if r["full_text"])
    print(f"    Vol 1 clauses with PDF text   : {matched_text}/{len(vol1)}")

    wb = Workbook()
    wb.remove(wb.active)

    # ---------------- Cover ----------------
    ws = wb.create_sheet("Cover", 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 80
    ws.column_dimensions['D'].width = 18

    ws.merge_cells("B2:D2"); c = ws["B2"]
    c.value = "369 ALLIANCE — Master NCC + Australian Standards Reference (v2)"
    c.font = font(NAVY, bold=True, size=22); c.alignment = LEFT
    ws.row_dimensions[2].height = 32

    ws.merge_cells("B3:D3"); c = ws["B3"]
    c.value = "Definitive cross-referenced legislation register — full clause text, longitudinal mapping, filtered views."
    c.font = font(GOLD, italic=True, size=12); c.alignment = LEFT

    ws.merge_cells("B4:D4"); c = ws["B4"]
    c.value = f"Compiled {datetime.now().strftime('%d %b %Y')} · 369 Alliance, Sydney NSW · v2 with PDF-extracted clause text and historical editions"
    c.font = font(GREY_HDR, size=10); c.alignment = LEFT

    sheets_meta = [
        ("01 NCC 2022 Vol 1 Clauses",       f"All {len(vol1)} clauses + full PDF text + AS / AS Clause relationship columns"),
        ("02 NCC 2022 Vol 2 (HP)",          f"{len(vol2)} HP clauses + AS / AS Clause relationship columns"),
        ("03 Performance Requirements",     "P-clauses only — with linked AS standards + AS clauses"),
        ("04 Deemed-to-Satisfy",            "D-clauses only — with linked AS standards + AS clauses"),
        ("05 Verification Methods",         "V-clauses only — with linked AS standards + AS clauses"),
        ("06 NCC-AS Relationship Matrix",   "Atomic NCC Clause ↔ AS Standard ↔ AS Clause ↔ Defect rows"),
        ("07 NCC 2019 Vol 1",               f"{len(ncc2019)} historical clauses (Amendment 1)"),
        ("08 NCC 2016 Vol 1",               f"{len(ncc2016)} historical clauses (Amendment 1)"),
        ("09 NCC 2015 Vol 1",               f"{len(ncc2015)} historical clauses (original BCA)"),
        ("10 Cross-Edition Mapping",        "NCC 2022 ↔ 2019 ↔ 2016 ↔ 2015 longitudinal lookup"),
        ("11 NSW BC Element Codes",         f"{len(short_codes)} regulated short codes — W/F/S/BE/ES taxonomy"),
        ("12 NSW Defect Register 2022",     f"{len(ncc_as_reg) or len(nsw_def_2022)} defects — canonical NCC↔AS source"),
        ("13 NSW Defect Register 2019",     f"{len(nsw_def_2019)} legacy NSW defect register entries"),
        ("14 AS Standards Master",          f"{len(refstd)} NCC-referenced standards"),
        ("15 AS in Use (categorised)",      f"{len(asuse)} active AS standards by discipline"),
        ("16 Topic-NCC-AS Crosswalk",       f"{len(cross)} topic-driven cross-references"),
        ("17 Master Defect Library",        f"{len(defects)} iAuditor defect items (full descriptions)"),
        ("18 Quick Lookup by Element",      "Element/regime → NCC + AS"),
        ("19 Reference Standards Index",    "AS index by code"),
        ("20 Section Colour Legend",        "Visual key"),
    ]

    ws.merge_cells("B6:D6"); c = ws["B6"]
    c.value = "WORKBOOK CONTENTS"
    c.font = font(WHITE, bold=True, size=12); c.fill = hdr_fill(NAVY); c.alignment = LEFT
    c.border = BORDER_ALL
    r = 7
    for name, desc in sheets_meta:
        ws.cell(r, 2, name).font = font(NAVY, bold=True)
        ws.cell(r, 3, desc).font = font(GREY_HDR)
        for col in (2, 3): ws.cell(r, col).alignment = LEFT
        r += 1

    # Sources
    r += 1
    ws.merge_cells(f"B{r}:D{r}"); c = ws.cell(r, 2)
    c.value = "AUTHORITATIVE SOURCES MERGED"
    c.font = font(WHITE, bold=True, size=12); c.fill = hdr_fill(NAVY); c.alignment = LEFT
    c.border = BORDER_ALL
    r += 1
    sources = [
        ("NCC 2022 Vol 1 PDF (full clause text)",       r"1. Legislation\NCC\PDF\NCC-2022-Volume-One-PCD.pdf"),
        ("NCC 2022 Vol 1 clause CSVs (82 files)",       r"1. Legislation\NCC\NCC_2022_Clauses_*.csv"),
        ("NCC 2022 Vol 2 HP all clauses",               r"1. Legislation\NCC\NCC_2022_HP_allClause.csv"),
        ("NCC 2019 Amendment 1 clause CSVs (50 files)", r"1. Legislation\NCC\NCC_2019A1_*.csv"),
        ("NCC 2016 Amendment 1 section CSVs",           r"1. Legislation\NCC\NCC2016A1 SECTION *.csv"),
        ("NCC 2015 section CSVs (original BCA)",        r"1. Legislation\NCC\NCC2015_Section*.csv"),
        ("NCC Vol 2 referenced standards",              r"1. Legislation\NCC\NCC Volume 2 Standards.csv"),
        ("NCC Building Reference Guide (2019/2020/2021)", r"1. Legislation\NCC Building Reference Guide-Updated with new links 21 October 2021.xlsx"),
        ("Australian Standards in use",                  r"1. Legislation\AS in use.docx"),
        ("Master Defect Library iAuditor",               r"1. Working folder\New Master_Defect Library iAuditor_r.xlsx"),
    ]
    for label, path in sources:
        ws.cell(r, 2, label).font = font(NAVY, bold=True)
        ws.cell(r, 3, path).font = font(GREY_HDR, italic=True, size=10)
        r += 1

    # Stats
    r += 1
    ws.merge_cells(f"B{r}:D{r}"); c = ws.cell(r, 2)
    c.value = "REGISTER STATISTICS"
    c.font = font(WHITE, bold=True, size=12); c.fill = hdr_fill(NAVY); c.alignment = LEFT
    c.border = BORDER_ALL
    r += 1
    stats = [
        ("NCC 2022 Vol 1 clauses",          len(vol1)),
        ("  with full PDF clause text",     matched_text),
        ("NCC 2022 Vol 2 (HP) clauses",     len(vol2)),
        ("NCC 2019 historical clauses",     len(ncc2019)),
        ("NCC 2016 historical clauses",     len(ncc2016)),
        ("NCC 2015 historical clauses",     len(ncc2015)),
        ("NSW BC element short codes",      len(short_codes)),
        ("NSW defect register (2022)",      len(nsw_def_2022)),
        ("NSW defect register (2019)",      len(nsw_def_2019)),
        ("Referenced standards (NCC2019+2022)", len(refstd)),
        ("AS in use (categorised)",         len(asuse)),
        ("Topic-AS crosswalk entries",      len(cross)),
        ("iAuditor defect library items",   len(defects)),
        ("GRAND TOTAL legislation rows",    len(vol1)+len(vol2)+len(refstd)+len(asuse)+len(cross)+len(defects)+len(ncc2019)+len(ncc2016)+len(ncc2015)+len(short_codes)+len(nsw_def_2022)+len(nsw_def_2019)),
    ]
    for label, n in stats:
        ws.cell(r, 2, label).font = font(NAVY, bold=label.strip().startswith("GRAND"))
        ws.cell(r, 3, n).font = font(AMBER, bold=True, size=12)
        ws.cell(r, 3).alignment = LEFT
        r += 1

    # ----- now build all data sheets in order -----
    s_vol1 = wb.create_sheet("01 NCC 2022 Vol 1 Clauses"); write_vol1_sheet(s_vol1, vol1, ncc_as_map)

    s_vol2 = wb.create_sheet("02 NCC 2022 Vol 2 (HP)")
    headers = ["NCC Version", "Volume", "Section #", "Section Title",
               "Part Code", "Part Title", "NCC Clause", "Clause Title",
               "AS (Australian Standard)", "AS Clause", "Linked Defect IDs", "Source"]
    write_header_row(s_vol2, headers, "4F6228", WHITE)
    for i, row in enumerate(vol2, 2):
        s_vol2.cell(i, 1, row["ncc_version"])
        s_vol2.cell(i, 2, row["volume"])
        sc = s_vol2.cell(i, 3, row["section"]); sc.alignment = CENTER
        sc.fill = hdr_fill(SECTION_COLOURS["V2"]); sc.font = font(WHITE, bold=True)
        s_vol2.cell(i, 4, row["section_title"]).font = font(NAVY, bold=True)
        s_vol2.cell(i, 5, row["part_code"]).alignment = CENTER
        s_vol2.cell(i, 6, row["part_title"])
        code = row["clause_code"]
        s_vol2.cell(i, 7, code).font = font(NAVY, bold=True)
        s_vol2.cell(i, 7).alignment = CENTER
        s_vol2.cell(i, 8, row["clause_title"])
        as_std, as_cl_pretty, defect_ids = aggregate_as_for_clause(ncc_as_map.get(code, []))
        ac = s_vol2.cell(i, 9, as_std)
        if as_std: ac.fill = hdr_fill("EFE3CC"); ac.font = font("8B5A3C", bold=True)
        ac.alignment = LEFTTOP
        s_vol2.cell(i, 10, as_cl_pretty).font = font(GREY_HDR, italic=True, size=9)
        s_vol2.cell(i, 10).alignment = LEFTTOP
        s_vol2.cell(i, 11, defect_ids).font = font(AMBER, bold=True, size=9)
        s_vol2.cell(i, 11).alignment = LEFTTOP
        s_vol2.cell(i, 12, row["source"]).font = font(GREY_HDR, size=9, italic=True)
        for col in range(1, len(headers)+1):
            s_vol2.cell(i, col).border = BORDER_ALL
            if col not in (3, 5, 7):
                s_vol2.cell(i, col).alignment = LEFTTOP
    autosize(s_vol2, max_widths={"D": 36, "F": 50, "H": 80, "I": 50, "J": 24,
                                 "K": 24, "L": 30, "B": 30})
    add_filter_and_freeze(s_vol2)

    # Filtered views
    s_p = wb.create_sheet("03 Performance Requirements")
    n_p = write_filtered_view(
        s_p, vol1, lambda r: r["clause_type"] == "Performance Requirement",
        "F4B084", "Performance Requirements", ncc_as_map)
    s_d = wb.create_sheet("04 Deemed-to-Satisfy")
    n_d = write_filtered_view(
        s_d, vol1, lambda r: r["clause_type"] == "Deemed-to-Satisfy",
        "C6EFCE", "Deemed-to-Satisfy", ncc_as_map)
    s_v = wb.create_sheet("05 Verification Methods")
    n_v = write_filtered_view(
        s_v, vol1, lambda r: r["clause_type"] == "Verification Method",
        "FFD966", "Verification Methods", ncc_as_map)
    print(f"    Filter views   P={n_p}  D={n_d}  V={n_v}")

    # NCC ↔ AS Relationship Matrix (one row per NCC clause × AS standard × AS clause)
    s_rel = wb.create_sheet("06 NCC-AS Relationship Matrix")
    headers = ["NCC Clause", "NCC Clause Title", "Section", "Clause Type",
               "AS (Australian Standard)", "AS Clause(s)",
               "Defect ID", "Brief Description"]
    write_header_row(s_rel, headers, "C07040", WHITE)
    # Build helper lookup: code -> (title, section, type) from vol1
    vol1_lookup = {r["clause_code"]: r for r in vol1 if r["clause_code"]}
    rownum = 2
    for code, entries in sorted(ncc_as_map.items()):
        for e in entries:
            ws_row = s_rel
            ws_row.cell(rownum, 1, code).font = font(NAVY, bold=True)
            ref = vol1_lookup.get(code, {})
            ws_row.cell(rownum, 2, ref.get("clause_title", ""))
            sec = ref.get("section", "")
            sc = ws_row.cell(rownum, 3, sec); sc.alignment = CENTER
            if sec:
                sc.fill = hdr_fill(SECTION_COLOURS.get(sec, "BFBFBF"))
                sc.font = font(WHITE, bold=True)
            ct_cell = ws_row.cell(rownum, 4, ref.get("clause_type", ""))
            c_col = CLAUSE_TYPE_COLOURS.get(ref.get("clause_type", ""))
            if c_col:
                ct_cell.fill = hdr_fill(c_col); ct_cell.alignment = CENTER
            as_cell = ws_row.cell(rownum, 5, e["as"])
            if e["as"]:
                as_cell.fill = hdr_fill("EFE3CC"); as_cell.font = font("8B5A3C", bold=True)
            ws_row.cell(rownum, 6, e["as_clauses"]).font = font(GREY_HDR, italic=True)
            ws_row.cell(rownum, 7, e["defect_id"]).font = font(AMBER, bold=True)
            ws_row.cell(rownum, 8, e["brief"])
            for col in range(1, len(headers)+1):
                ws_row.cell(rownum, col).border = BORDER_ALL
                if col not in (3, 4):
                    ws_row.cell(rownum, col).alignment = LEFTTOP
            rownum += 1
    autosize(s_rel, max_widths={"A": 14, "B": 50, "C": 6, "D": 22,
                                "E": 30, "F": 36, "G": 14, "H": 70})
    add_filter_and_freeze(s_rel)
    print(f"    NCC-AS relationship rows      : {rownum-2}")

    # Historical
    s_2019 = wb.create_sheet("07 NCC 2019 Vol 1");      write_historical_sheet(s_2019, ncc2019, "NCC 2019",      "5B6C8C")
    s_2016 = wb.create_sheet("08 NCC 2016 Vol 1");      write_historical_sheet(s_2016, ncc2016, "NCC 2016 (A1)", "6B7280")
    s_2015 = wb.create_sheet("09 NCC 2015 Vol 1");      write_historical_sheet(s_2015, ncc2015, "NCC 2015",      "7C8794")

    # Cross-edition
    s_cx = wb.create_sheet("10 Cross-Edition Mapping")
    n_cx = write_cross_edition(s_cx, vol1, ncc2019, ncc2016, ncc2015)
    print(f"    Cross-edition mappings        : {n_cx}")

    # ---------------- 11 NSW BC Element Codes ----------------
    s_sc = wb.create_sheet("11 NSW BC Element Codes")
    headers = ["Element Code", "Regime / Element", "Description", "Visual Code"]
    write_header_row(s_sc, headers, "1A1A2E", WHITE)
    elem_palette = {
        "Waterproofing":      "1E90FF",
        "Fire Safety":        "B22222",
        "Structure":          "8B5A3C",
        "Building Enclosure": "DAA520",
        "Essential Service":  "20B2AA",
    }
    for i, row in enumerate(short_codes, 2):
        cell = s_sc.cell(i, 1, row["code"]); cell.font = font(NAVY, bold=True); cell.alignment = CENTER
        elem = row["element"]
        ec = s_sc.cell(i, 2, elem); ec.alignment = CENTER
        colour = elem_palette.get(elem)
        if colour:
            ec.fill = hdr_fill(colour); ec.font = font(WHITE, bold=True)
        s_sc.cell(i, 3, row["description"]).alignment = LEFTTOP
        # Visual code chip
        chip = s_sc.cell(i, 4, "")
        if colour: chip.fill = hdr_fill(colour)
        for col in range(1, len(headers)+1):
            s_sc.cell(i, col).border = BORDER_ALL
    autosize(s_sc, max_widths={"A": 14, "B": 24, "C": 60, "D": 12})
    add_filter_and_freeze(s_sc)

    # ---------------- 12 NSW Defect Register 2022 (rich NCC and AS file) ----------------
    s_d22 = wb.create_sheet("12 NSW Defect Register 2022")
    register_source = ncc_as_reg if ncc_as_reg else nsw_def_2022
    if register_source:
        # Use canonical column order
        all_headers = list(register_source[0].keys())
        # Preferred ordering
        preferred = ["Link", "Element Code", "Item", "Version", "Unique ID",
                     "Brief Description/What is the defect?", "Consequence",
                     "NCC Quotes", "AS Used", "AS Clauses",
                     "Reasons why it is a serious defects", "Scope of works",
                     "Available"]
        headers = [h for h in preferred if h in all_headers] + [h for h in all_headers if h not in preferred]
        write_header_row(s_d22, headers, "8B5A3C", WHITE)
        code_regime = {sc["code"]: sc["element"] for sc in short_codes}
        ec_idx       = headers.index("Element Code")+1 if "Element Code" in headers else None
        nccq_idx     = headers.index("NCC Quotes")+1   if "NCC Quotes" in headers else None
        as_used_idx  = headers.index("AS Used")+1      if "AS Used" in headers else None
        as_cl_idx    = headers.index("AS Clauses")+1   if "AS Clauses" in headers else None
        uid_idx      = headers.index("Unique ID")+1    if "Unique ID" in headers else None
        for i, rec in enumerate(register_source, 2):
            for c, h in enumerate(headers, 1):
                cell = s_d22.cell(i, c, rec.get(h, ""))
                cell.alignment = LEFTTOP
                cell.border = BORDER_ALL
            if ec_idx:
                ec_val = rec.get("Element Code") or ""
                regime = code_regime.get(ec_val, "")
                colour = elem_palette.get(regime)
                if colour:
                    cell = s_d22.cell(i, ec_idx)
                    cell.fill = hdr_fill(colour); cell.font = font(WHITE, bold=True)
                    cell.alignment = CENTER
            if uid_idx:
                s_d22.cell(i, uid_idx).font = font(NAVY, bold=True)
                s_d22.cell(i, uid_idx).alignment = CENTER
            if nccq_idx:
                s_d22.cell(i, nccq_idx).font = font(AMBER, bold=True)
            if as_used_idx:
                s_d22.cell(i, as_used_idx).font = font("8B5A3C", bold=True)
            if as_cl_idx:
                s_d22.cell(i, as_cl_idx).font = font(GREY_HDR, italic=True)
        autosize(s_d22, max_widths={"A": 10, "B": 8, "C": 8, "D": 8, "E": 12,
                                     "F": 60, "G": 60, "H": 30, "I": 26, "J": 24,
                                     "K": 60, "L": 60, "M": 10})
        add_filter_and_freeze(s_d22)

    # ---------------- 13 NSW Defect Register 2019 ----------------
    s_d19 = wb.create_sheet("13 NSW Defect Register 2019")
    if nsw_def_2019:
        headers = list(nsw_def_2019[0].keys())
        write_header_row(s_d19, headers, "5B6C8C", WHITE)
        code_regime = {sc["code"]: sc["element"] for sc in short_codes}
        ec_idx  = headers.index("Element Code")+1 if "Element Code" in headers else None
        uid_idx = headers.index("Unique ID")+1   if "Unique ID" in headers else None
        for i, rec in enumerate(nsw_def_2019, 2):
            for c, h in enumerate(headers, 1):
                cell = s_d19.cell(i, c, rec.get(h, ""))
                cell.alignment = LEFTTOP
                cell.border = BORDER_ALL
            if ec_idx:
                ec_val = rec.get("Element Code") or ""
                regime = code_regime.get(ec_val, "")
                colour = elem_palette.get(regime)
                if colour:
                    cell = s_d19.cell(i, ec_idx)
                    cell.fill = hdr_fill(colour); cell.font = font(WHITE, bold=True)
                    cell.alignment = CENTER
            if uid_idx:
                s_d19.cell(i, uid_idx).font = font(NAVY, bold=True)
                s_d19.cell(i, uid_idx).alignment = CENTER
        autosize(s_d19, max_widths={"A": 10, "B": 8, "C": 12, "D": 8, "E": 10, "F": 14,
                                     "G": 60, "H": 60})
        add_filter_and_freeze(s_d19)

    # Standards
    s_std = wb.create_sheet("14 AS Standards Master")
    headers = ["NCC Version", "Standard Code", "Year", "Title", "Version / Notes", "Source"]
    write_header_row(s_std, headers, "8B5A3C", WHITE)
    for i, row in enumerate(refstd, 2):
        v = row["ncc_version"]; vc = s_std.cell(i, 1, v); vc.alignment = CENTER
        if   v == "NCC2022": vc.fill = hdr_fill("C6EFCE"); vc.font = font(NAVY, bold=True)
        elif v == "NCC2019": vc.fill = hdr_fill("FFE699"); vc.font = font(NAVY, bold=True)
        s_std.cell(i, 2, row["code"]).font = font(NAVY, bold=True)
        s_std.cell(i, 3, row["year"]).alignment = CENTER
        s_std.cell(i, 4, row["title"]).alignment = LEFTTOP
        s_std.cell(i, 5, row["version_note"]).font = font(GREY_HDR, italic=True)
        s_std.cell(i, 6, row["source"]).font = font(GREY_HDR, size=9)
        for col in range(1, len(headers)+1):
            s_std.cell(i, col).border = BORDER_ALL
    autosize(s_std, max_widths={"D": 70, "E": 30, "F": 32, "B": 32})
    add_filter_and_freeze(s_std)

    s_asu = wb.create_sheet("15 AS in Use (categorised)")
    headers = ["Discipline / Category", "AS Code", "Title"]
    write_header_row(s_asu, headers, GOLD, WHITE)
    cat_colours = {"Waterproofing": "1E90FF", "Fire": "B22222", "Structure": "8B5A3C",
                   "Wind Classification": "20B2AA", "Cladding": "C07040",
                   "Essential Services": "DAA520"}
    for i, row in enumerate(asuse, 2):
        c = s_asu.cell(i, 1, row["category"]); c.alignment = CENTER
        if row["category"] in cat_colours:
            c.fill = hdr_fill(cat_colours[row["category"]]); c.font = font(WHITE, bold=True)
        s_asu.cell(i, 2, row["code"]).font = font(NAVY, bold=True)
        s_asu.cell(i, 3, row["title"]).alignment = LEFTTOP
        for col in range(1, len(headers)+1):
            s_asu.cell(i, col).border = BORDER_ALL
    autosize(s_asu, max_widths={"A": 28, "B": 26, "C": 90})
    add_filter_and_freeze(s_asu)

    s_cw = wb.create_sheet("16 Topic-NCC-AS Crosswalk")
    headers = ["Topic / Family", "Complaint Item / Sub-issue",
               "BCA Vol 1 (2019)", "BCA Vol 2 (2019)", "BCA Vol 3 (2019)",
               "Australian Standard", "Guide to Standards & Tolerances", "Other", "Previous Versions"]
    write_header_row(s_cw, headers, AMBER, WHITE)
    for i, row in enumerate(cross, 2):
        s_cw.cell(i, 1, row["topic"]).font = font(NAVY, bold=True)
        s_cw.cell(i, 2, row["complaint_item"])
        s_cw.cell(i, 3, row["ncc_vol1_2019"])
        s_cw.cell(i, 4, row["ncc_vol2_2019"])
        s_cw.cell(i, 5, row["ncc_vol3_2019"])
        s_cw.cell(i, 6, row["as_reference"])
        s_cw.cell(i, 7, row["guide_to_std"])
        s_cw.cell(i, 8, row["other"])
        s_cw.cell(i, 9, row["previous"]).font = font(GREY_HDR, italic=True, size=9)
        for col in range(1, len(headers)+1):
            s_cw.cell(i, col).border = BORDER_ALL
            s_cw.cell(i, col).alignment = LEFTTOP
    autosize(s_cw, max_widths={"A": 26, "B": 50, "C": 45, "D": 45, "E": 30, "F": 45, "G": 35, "H": 25, "I": 30})
    add_filter_and_freeze(s_cw)

    s_def = wb.create_sheet("17 Master Defect Library")
    if defects:
        headers = list(defects[0].keys())
        write_header_row(s_def, headers, "8B5A3C", WHITE)
        for i, rec in enumerate(defects, 2):
            for c, h in enumerate(headers, 1):
                cell = s_def.cell(i, c, rec.get(h, ""))
                cell.alignment = LEFTTOP; cell.border = BORDER_ALL
            did = (rec.get("Defect ID") or "").strip()
            be  = (rec.get("Building element") or "").strip()
            cidx = headers.index("Defect ID") + 1 if "Defect ID" in headers else None
            beidx= headers.index("Building element") + 1 if "Building element" in headers else None
            if cidx:
                code_cell = s_def.cell(i, cidx)
                colour = REGIME_COLOURS.get(be)
                if not colour and did:
                    code_letter = re.match(r"^[A-Z]+", did)
                    if code_letter:
                        colour = {"W": "1E90FF", "F": "B22222", "S": "8B5A3C",
                                  "BE": "DAA520", "ES": "20B2AA"}.get(code_letter.group(0))
                if colour:
                    code_cell.fill = hdr_fill(colour); code_cell.font = font(WHITE, bold=True)
                    code_cell.alignment = CENTER
            if beidx:
                be_cell = s_def.cell(i, beidx); colour = REGIME_COLOURS.get(be)
                if colour:
                    be_cell.fill = hdr_fill(colour); be_cell.font = font(WHITE, bold=True)
                    be_cell.alignment = CENTER
        autosize(s_def, max_widths={"A": 14, "B": 26, "C": 28, "D": 60, "E": 60, "F": 50,
                                    "G": 50, "H": 50, "I": 50, "J": 18, "K": 50, "L": 35, "M": 16, "N": 30})
        add_filter_and_freeze(s_def)

    s_qk = wb.create_sheet("18 Quick Lookup by Element")
    headers = ["Building Element / Regime", "Sub-Category", "Defect Codes",
               "Indicative NCC 2022 Refs", "Indicative AS Refs", "Defect Count"]
    write_header_row(s_qk, headers, "1E90FF", WHITE)
    by_element = {}
    for d in defects:
        be = (d.get("Building element") or "Unspecified").strip()
        sc = (d.get("Sub-categories") or "").strip()
        e = by_element.setdefault((be, sc), {"codes": [], "ncc22": set(), "as_refs": set()})
        e["codes"].append((d.get("Defect ID") or "").strip())
        for src in (d.get("NCC 2019 Reference"), d.get("NCC 2022 Reference"),
                    d.get("Requirement for standard of work"), d.get("Reason why it is a serious defect")):
            if isinstance(src, str):
                for as_code in re.findall(r"AS(?:/NZS)?\s*\d+(?:\.\d+)*(?::\d{4})?", src):
                    e["as_refs"].add(as_code.strip())
        ncc22 = d.get("NCC 2022 Reference") or ""
        if isinstance(ncc22, str):
            for token in re.findall(r"\b(?:Volume\s*\w+|Section\s*\w+|Part\s*\w+|Specification\s*\w+|Clause\s*\w+|\d+\.\d+(?:\.\d+)*|[A-Z]+\d+[A-Z]?\d*)\b", ncc22):
                e["ncc22"].add(token)
    rownum = 2
    for (be, sc), v in sorted(by_element.items(), key=lambda kv: (kv[0][0], kv[0][1])):
        bcell = s_qk.cell(rownum, 1, be); colour = REGIME_COLOURS.get(be)
        if colour:
            bcell.fill = hdr_fill(colour); bcell.font = font(WHITE, bold=True); bcell.alignment = CENTER
        else:
            bcell.font = font(NAVY, bold=True)
        s_qk.cell(rownum, 2, sc)
        s_qk.cell(rownum, 3, ", ".join(sorted(set(c for c in v["codes"] if c))))
        s_qk.cell(rownum, 4, ", ".join(sorted(v["ncc22"])))
        s_qk.cell(rownum, 5, ", ".join(sorted(v["as_refs"])))
        s_qk.cell(rownum, 6, len([c for c in v["codes"] if c])).alignment = CENTER
        for col in range(1, len(headers)+1):
            s_qk.cell(rownum, col).border = BORDER_ALL
        rownum += 1
    autosize(s_qk, max_widths={"A": 32, "B": 28, "C": 50, "D": 50, "E": 40, "F": 14})
    add_filter_and_freeze(s_qk)

    s_ix = wb.create_sheet("19 Reference Standards Index")
    headers = ["Standard Code", "Title", "NCC Version(s)", "Year(s)"]
    write_header_row(s_ix, headers, GOLD, WHITE)
    by_code = {}
    for r in refstd:
        if not r["code"].strip(): continue
        d = by_code.setdefault(r["code"].strip(), {"titles": set(), "vers": set(), "years": set()})
        if r["title"]: d["titles"].add(r["title"])
        if r["ncc_version"]: d["vers"].add(r["ncc_version"])
        if r["year"]: d["years"].add(r["year"])
    for r in asuse:
        if not r["code"].strip(): continue
        d = by_code.setdefault(r["code"].strip(), {"titles": set(), "vers": set(), "years": set()})
        if r["title"]: d["titles"].add(r["title"])
        d["vers"].add("AS in use list")
    for i, (code, d) in enumerate(sorted(by_code.items()), 2):
        s_ix.cell(i, 1, code).font = font(NAVY, bold=True)
        s_ix.cell(i, 2, "; ".join(sorted(d["titles"]))).alignment = LEFTTOP
        vers_str = ", ".join(sorted(d["vers"]))
        vc = s_ix.cell(i, 3, vers_str); vc.alignment = CENTER
        if "NCC2022" in vers_str and "NCC2019" in vers_str: vc.fill = hdr_fill("C6EFCE")
        elif "NCC2022" in vers_str: vc.fill = hdr_fill("EFE3CC")
        elif "NCC2019" in vers_str: vc.fill = hdr_fill("FFE699")
        s_ix.cell(i, 4, ", ".join(sorted(d["years"]))).alignment = CENTER
        for col in range(1, len(headers)+1):
            s_ix.cell(i, col).border = BORDER_ALL
    autosize(s_ix, max_widths={"A": 30, "B": 80, "C": 22, "D": 16})
    add_filter_and_freeze(s_ix)

    # Legend
    s_lg = wb.create_sheet("20 Section Colour Legend")
    s_lg.sheet_view.showGridLines = False
    s_lg.column_dimensions['A'].width = 3
    s_lg.column_dimensions['B'].width = 20
    s_lg.column_dimensions['C'].width = 38
    s_lg.column_dimensions['D'].width = 16

    s_lg.merge_cells("B2:D2"); h = s_lg["B2"]
    h.value = "VISUAL LEGEND — NCC Sections, Clause Types, Regimes"
    h.font = font(NAVY, bold=True, size=16); h.alignment = LEFT

    legend_blocks = [
        ("NCC VOLUME ONE — SECTIONS", [
            ("A — General provisions",         "5B6C8C"),
            ("B — Structure",                  "8B5A3C"),
            ("C — Fire resistance",            "B22222"),
            ("D — Access & egress",            "FF8C00"),
            ("E — Services & equipment",       "DAA520"),
            ("F — Health & amenity",           "1E90FF"),
            ("G — Ancillary provisions",       "228B22"),
            ("I — Special use buildings",      "9370DB"),
            ("J — Energy efficiency",          "20B2AA"),
            ("S — Specifications",             "696969"),
        ]),
        ("NCC VOLUME TWO — HOUSING PROVISIONS", [
            ("Vol 2 (HP) Sections 2-13",       "4F6228"),
        ]),
        ("CLAUSE TYPE", [
            ("Objective (O)",                  "B0C4DE"),
            ("Functional Statement (F)",       "C2E0FF"),
            ("Performance Requirement (P)",    "F4B084"),
            ("Verification Method (V)",        "FFD966"),
            ("Deemed-to-Satisfy (D)",          "C6EFCE"),
            ("Specification Clause (Sx C)",    "D9D9D9"),
            ("Governing / General (G)",        "E7E6E6"),
        ]),
        ("DEFECT REGIMES (NSW Building Commissioner)", [
            ("W — Waterproofing",              "1E90FF"),
            ("F — Fire safety",                "B22222"),
            ("S — Structure",                  "8B5A3C"),
            ("BE — Building Enclosure",        "DAA520"),
            ("ES — Essential Services",        "20B2AA"),
        ]),
        ("NCC VERSION HIGHLIGHT", [
            ("Standard referenced by NCC 2022", "C6EFCE"),
            ("Standard referenced by NCC 2019", "FFE699"),
            ("Referenced in BOTH editions",     "EFE3CC"),
        ]),
    ]
    r = 4
    for title, items in legend_blocks:
        s_lg.merge_cells(f"B{r}:D{r}")
        c = s_lg.cell(r, 2, title); c.fill = hdr_fill(NAVY)
        c.font = font(WHITE, bold=True, size=11); c.alignment = LEFT
        r += 1
        for label, hex_ in items:
            s_lg.cell(r, 2, label).font = font(NAVY)
            sw = s_lg.cell(r, 3, ""); sw.fill = hdr_fill(hex_)
            s_lg.cell(r, 4, "#" + hex_).font = font(GREY_HDR, italic=True, size=9)
            s_lg.cell(r, 4).alignment = CENTER
            r += 1
        r += 1

    # Cover hyperlinks (after sheets exist)
    cover = wb["Cover"]
    sheet_link_map = [
        (7,  "01 NCC 2022 Vol 1 Clauses"),
        (8,  "02 NCC 2022 Vol 2 (HP)"),
        (9,  "03 Performance Requirements"),
        (10, "04 Deemed-to-Satisfy"),
        (11, "05 Verification Methods"),
        (12, "06 NCC-AS Relationship Matrix"),
        (13, "07 NCC 2019 Vol 1"),
        (14, "08 NCC 2016 Vol 1"),
        (15, "09 NCC 2015 Vol 1"),
        (16, "10 Cross-Edition Mapping"),
        (17, "11 NSW BC Element Codes"),
        (18, "12 NSW Defect Register 2022"),
        (19, "13 NSW Defect Register 2019"),
        (20, "14 AS Standards Master"),
        (21, "15 AS in Use (categorised)"),
        (22, "16 Topic-NCC-AS Crosswalk"),
        (23, "17 Master Defect Library"),
        (24, "18 Quick Lookup by Element"),
        (25, "19 Reference Standards Index"),
        (26, "20 Section Colour Legend"),
    ]
    for row_num, sheet_name in sheet_link_map:
        cell = cover.cell(row_num, 2)
        cell.hyperlink = f"#'{sheet_name}'!A1"
        cell.font = font("0563C1", bold=True, size=11)

    print("[ ] Saving …")
    try:
        wb.save(OUT)
    except PermissionError:
        # If user has the file open, save with timestamp suffix
        alt = OUT.with_name(OUT.stem + f"_{datetime.now():%H%M%S}" + OUT.suffix)
        wb.save(alt)
        print(f"[!] Original file locked — saved as {alt.name}")
        return alt
    sz = OUT.stat().st_size / 1024
    print(f"[OK] {OUT}  ({sz:,.0f} KB)")
    return OUT

if __name__ == "__main__":
    main()
