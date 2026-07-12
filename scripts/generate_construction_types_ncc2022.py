# -*- coding: utf-8 -*-
"""Generate NCC 2022 Types A, B, C construction scheme workbook.

Builds an Excel workbook with:
  - Cover sheet referencing NCC 2022 Volume 1 Part C2 + Spec 5
  - Table C2D2 (Type of construction required by Class & Rise in Storeys)
  - Three visual design schemes (Type A / Type B / Type C)
  - FRL summary tables from Specification 5
  - Fire safety systems matrix per construction type
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Brand palette (369 Alliance) and helpers
# ---------------------------------------------------------------------------
NAVY = "1A1A2E"
GOLD = "A68A64"
AMBER = "C07040"
CREAM = "F5EFE6"
LIGHT_GREY = "EFEFEF"
WHITE = "FFFFFF"
RED = "C0392B"
GREEN = "2E7D32"
ORANGE = "E67E22"
SKY = "BBD9E8"

THIN = Side(border_style="thin", color="555555")
MED = Side(border_style="medium", color=NAVY)
ALL_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
ALL_MED = Border(left=MED, right=MED, top=MED, bottom=MED)


def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def title_font(size=14, color=WHITE, bold=True):
    return Font(name="Calibri", size=size, bold=bold, color=color)


def body_font(size=10, color="111111", bold=False):
    return Font(name="Calibri", size=size, bold=bold, color=color)


def write_header(ws, cell_range, text, bg=NAVY, color=WHITE, size=14):
    ws.merge_cells(cell_range)
    top_left = cell_range.split(":")[0]
    c = ws[top_left]
    c.value = text
    c.fill = fill(bg)
    c.font = title_font(size=size, color=color)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def set_col_widths(ws, widths):
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w


def set_row_heights(ws, heights):
    for idx, h in enumerate(heights, start=1):
        ws.row_dimensions[idx].height = h


# ---------------------------------------------------------------------------
# Sheet 1 — Cover / Index
# ---------------------------------------------------------------------------
def build_cover(wb):
    ws = wb.active
    ws.title = "Cover"
    set_col_widths(ws, [3, 28, 70, 3])

    ws.row_dimensions[2].height = 40
    write_header(
        ws,
        "B2:C2",
        "NCC 2022 — Types of Construction (A, B, C)\nDesign Schemes & Fire Resistance Reference",
        bg=NAVY,
        size=16,
    )

    ws["B4"] = "Reference framework"
    ws["B4"].font = body_font(11, color=NAVY, bold=True)
    ws["C4"] = (
        "National Construction Code 2022 — Volume One, Building Code of Australia "
        "(BCA), Part C2 (Type of construction) and Specification 5 (Fire-resisting "
        "construction). Cross-referenced with Part D2/D3 (Access & egress), Part E1 "
        "(Fire fighting equipment), E2 (Smoke hazard management) and E3 (Lift "
        "installations)."
    )
    ws["C4"].alignment = Alignment(wrap_text=True, vertical="top")

    rows = [
        ("Index", ""),
        ("Sheet 1 — Cover", "This sheet."),
        (
            "Sheet 2 — Table C2D2",
            "Type of construction required by Class of building and Rise in Storeys (NCC 2022 C2D2).",
        ),
        (
            "Sheet 3 — Type A scheme",
            "Most fire-resistant. Diagram + design example for a Class 2 / 3 / 9 building of 4+ storeys, "
            "or Class 5 / 6 / 7 / 8 of 3+ storeys.",
        ),
        (
            "Sheet 4 — Type B scheme",
            "Intermediate. Diagram + design example for a 2-storey Class 2/3/9 or 3-storey Class 5-8.",
        ),
        (
            "Sheet 5 — Type C scheme",
            "Least restrictive. Diagram + design example for low-rise / single-storey buildings.",
        ),
        (
            "Sheet 6 — FRL summary",
            "Fire Resistance Levels for each construction type from Specification 5, Tables 3, 4 and 5.",
        ),
        (
            "Sheet 7 — Fire safety systems",
            "Required fire safety systems per construction type and effective height.",
        ),
        (
            "Sheet 8 — Compartmentation & exits",
            "Maximum floor areas / volumes and minimum number of exits per Class and type.",
        ),
    ]

    r = 6
    for k, v in rows:
        ws.cell(row=r, column=2, value=k).font = body_font(10, color=NAVY, bold=True)
        c = ws.cell(row=r, column=3, value=v)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        c.font = body_font(10)
        ws.row_dimensions[r].height = 30 if v else 18
        r += 1

    ws["B" + str(r + 1)] = "Notes"
    ws["B" + str(r + 1)].font = body_font(11, color=AMBER, bold=True)
    notes = (
        "1. Clause numbers follow the NCC 2022 Adoption (1 May 2023) referencing "
        "format (e.g. C2D2, E1D2). For projects on transitional pathways, use the "
        "NCC 2019 A3.0 numbering equivalent.\n"
        "2. Performance-based alternatives are permitted under the NCC Performance "
        "Solution pathway (A2G2 / A2G4) and are not shown on this scheme.\n"
        "3. Diagrams are conceptual and not to scale. Always read in conjunction "
        "with the full NCC text and any DTS Specification referenced."
    )
    ws.cell(row=r + 2, column=2, value=notes).alignment = Alignment(
        wrap_text=True, vertical="top"
    )
    ws.merge_cells(start_row=r + 2, start_column=2, end_row=r + 6, end_column=3)


# ---------------------------------------------------------------------------
# Sheet 2 — Table C2D2
# ---------------------------------------------------------------------------
def build_c2d2(wb):
    ws = wb.create_sheet("Table C2D2")
    set_col_widths(ws, [3, 22, 18, 18, 18, 18, 22, 3])

    write_header(
        ws,
        "B2:G2",
        "NCC 2022 Table C2D2 — Type of construction required",
        size=14,
    )

    ws["B4"] = (
        "The Type of construction (A, B or C) is determined from the Rise in Storeys "
        "(measured under C2D4) and the Class of the building under Part A6. Type A is "
        "the most fire-resistant; Type C the least."
    )
    ws.merge_cells("B4:G4")
    ws["B4"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[4].height = 38

    headers = [
        "Rise in Storeys",
        "Class 2",
        "Class 3",
        "Class 9 (a, b, c)",
        "Class 5, 6, 7, 8",
        "Notes",
    ]
    for i, h in enumerate(headers, start=2):
        c = ws.cell(row=6, column=i, value=h)
        c.fill = fill(GOLD)
        c.font = title_font(11, color=WHITE)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = ALL_THIN
    ws.row_dimensions[6].height = 32

    table = [
        ("4 or more", "A", "A", "A", "A", "Class 2/3 over 25 m must also comply with E1D7 (sprinklers)."),
        ("3", "A", "A", "A", "B", "Class 9b assembly often Type A regardless — check C2D2."),
        ("2", "B", "B", "B", "C", "Class 2/3 of 2 storeys may use Type C if total floor area ≤ 500 m² (C2D2(b))."),
        ("1", "C", "C", "C", "C", "Class 9c (aged care) — special concession; check C2D2(c)."),
    ]

    type_colors = {"A": RED, "B": ORANGE, "C": GREEN}

    for r_idx, row in enumerate(table, start=7):
        for c_idx, val in enumerate(row, start=2):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = ALL_THIN
            cell.alignment = Alignment(
                horizontal="center" if c_idx <= 6 else "left",
                vertical="center",
                wrap_text=True,
            )
            if c_idx in (3, 4, 5, 6) and val in type_colors:
                cell.fill = fill(type_colors[val])
                cell.font = title_font(11, color=WHITE)
            else:
                cell.font = body_font(10)
        ws.row_dimensions[r_idx].height = 28

    ws["B12"] = "Legend"
    ws["B12"].font = body_font(11, color=NAVY, bold=True)
    legend = [
        ("Type A", RED, "Most fire-resistant. FRL up to 240/240/240."),
        ("Type B", ORANGE, "Intermediate. FRL typically 120/120/120."),
        ("Type C", GREEN, "Least restrictive. FRL 60/60/60 or non-combustible only near fire source feature."),
    ]
    for i, (name, color, desc) in enumerate(legend, start=13):
        a = ws.cell(row=i, column=2, value=name)
        a.fill = fill(color)
        a.font = title_font(11, color=WHITE)
        a.alignment = Alignment(horizontal="center", vertical="center")
        a.border = ALL_THIN
        b = ws.cell(row=i, column=3, value=desc)
        b.font = body_font(10)
        b.alignment = Alignment(wrap_text=True, vertical="center")
        b.border = ALL_THIN
        ws.merge_cells(start_row=i, start_column=3, end_row=i, end_column=7)
        ws.row_dimensions[i].height = 22


# ---------------------------------------------------------------------------
# Helpers for the Type A/B/C scheme sheets
# ---------------------------------------------------------------------------
def style_level_cell(cell, label, color):
    cell.value = label
    cell.fill = fill(color)
    cell.font = body_font(10, color=NAVY, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = ALL_THIN


def style_facade_cell(cell, color):
    cell.fill = fill(color)
    cell.border = ALL_THIN


def write_callout(ws, row, col, text, color=GOLD, span=4):
    ws.merge_cells(
        start_row=row, start_column=col, end_row=row, end_column=col + span - 1
    )
    c = ws.cell(row=row, column=col, value=text)
    c.fill = fill(color)
    c.font = body_font(9, color=NAVY)
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
    c.border = ALL_THIN


# ---------------------------------------------------------------------------
# Sheet 3 — Type A scheme (8-storey Class 2 example)
# ---------------------------------------------------------------------------
def build_type_a(wb):
    ws = wb.create_sheet("Type A scheme")
    set_col_widths(ws, [3, 6, 8, 8, 8, 8, 18, 26, 26, 3])

    write_header(
        ws,
        "B2:I2",
        "Type A — Most fire-resistant (NCC 2022 C2D2 + Spec 5 Table 3)",
        bg=RED,
        size=14,
    )

    ws["B4"] = (
        "Design example: Class 2 residential apartment building, 8 storeys above "
        "ground, basement carpark (Class 7a), effective height ≈ 26 m. Triggers "
        "Type A construction (4+ storeys Class 2) and the >25 m fire safety package."
    )
    ws.merge_cells("B4:I4")
    ws["B4"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[4].height = 36

    levels = [
        ("L8", "Class 2 — Apartments", "Roof / plant"),
        ("L7", "Class 2 — Apartments", "Effective height >25 m → E1D7 dual water, E3D9 emergency lift"),
        ("L6", "Class 2 — Apartments", ""),
        ("L5", "Class 2 — Apartments", ""),
        ("L4", "Class 2 — Apartments", ""),
        ("L3", "Class 2 — Apartments", "Hydrants required from L3+ during construction (E1D2)"),
        ("L2", "Class 2 — Apartments", ""),
        ("L1", "Class 2 — Apartments + lobby", ""),
        ("GF", "Lobby + Class 6 retail tenancy", "Fire Control Room at access level (E1D14)"),
        ("B1", "Class 7a — Carpark + services", "Sprinklered (Spec E1.5); smoke exhaust per E2D14"),
    ]

    start_row = 6
    for i, (lbl, occ, note) in enumerate(levels):
        row = start_row + i
        # Left wall
        style_facade_cell(ws.cell(row=row, column=2), NAVY)
        # Storey label
        style_level_cell(ws.cell(row=row, column=3), lbl, CREAM)
        # Occupancy
        c = ws.cell(row=row, column=4, value=occ)
        c.fill = fill(LIGHT_GREY)
        c.font = body_font(9)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = ALL_THIN
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
        # Right wall
        style_facade_cell(ws.cell(row=row, column=7), NAVY)
        # Note
        n = ws.cell(row=row, column=8, value=note)
        n.font = body_font(9, color=AMBER)
        n.alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells(start_row=row, start_column=8, end_row=row, end_column=9)
        ws.row_dimensions[row].height = 28

    # Ground line
    ground_row = start_row + len(levels)
    for col in range(2, 8):
        ws.cell(row=ground_row, column=col).fill = fill("8B7355")
        ws.cell(row=ground_row, column=col).border = ALL_THIN
    ws.row_dimensions[ground_row].height = 6

    # Mandatory systems block
    sys_row = ground_row + 2
    write_header(
        ws,
        f"B{sys_row}:I{sys_row}",
        "Mandatory fire safety package — Type A, effective height >25 m",
        bg=AMBER,
        size=12,
    )

    items = [
        (
            "Fire isolated stairs",
            "Two required, fire-isolated, with stair pressurisation (AS 1668.1) or open access ramp.",
            "D2D6, D3D8, E2D14",
        ),
        (
            "Fire hydrant system",
            "Throughout, with booster connection at access level. Operational hydrants required during construction once height >12 m.",
            "E1D2, AS 2419.1",
        ),
        (
            "Fire hose reels",
            "From level 3 to top, coverage entire floor (≤ 4 m of any point).",
            "E1D3, AS 2441",
        ),
        (
            "Booster connections",
            "At Fire Brigade access point, IP rated, signposted.",
            "E1D2 + AS 2419.1 §6",
        ),
        (
            "Fire Control Room (FCR)",
            "Required when total floor area >500 m² and effective height >25 m. Room separated by 120/120/120.",
            "E1D14, Spec 20",
        ),
        (
            "Sprinkler system",
            "Class 2/3 over 25 m: dual water supply, fast-response heads, FIP at FCR.",
            "E1D6, E1D7, AS 2118.1",
        ),
        (
            "Emergency lift",
            "Required where effective height >25 m. Independent power, fire service control switch.",
            "E3D9, AS 1735.1",
        ),
        (
            "Smoke hazard management",
            "Apartments — smoke alarms AS 3786 + corridor smoke detectors. Carpark — mechanical smoke exhaust.",
            "E2D2, E2D5, Spec E2.2",
        ),
        (
            "Compartmentation",
            "Apartments separated by 90/90/90 sole-occupancy walls; floor 90/90/90; max compartment area 8 000 m².",
            "C3D3, Spec 5 Table 3",
        ),
        (
            "Two exits per storey",
            "In addition to any horizontal exit (D2D2 + D2D4).",
            "D2D2",
        ),
    ]

    hdr = ["System / requirement", "Description", "NCC 2022 reference"]
    for i, h in enumerate(hdr, start=2):
        c = ws.cell(row=sys_row + 2, column=i + (0 if i == 2 else (1 if i == 3 else 2)))
    # Manual placement to keep alignment clean
    headings_row = sys_row + 2
    ws.cell(row=headings_row, column=2, value="System / requirement").font = title_font(
        11, color=WHITE
    )
    ws.cell(row=headings_row, column=2).fill = fill(NAVY)
    ws.cell(row=headings_row, column=2).alignment = Alignment(
        horizontal="center", vertical="center"
    )
    ws.merge_cells(
        start_row=headings_row, start_column=2, end_row=headings_row, end_column=3
    )
    ws.cell(row=headings_row, column=4, value="Description").font = title_font(
        11, color=WHITE
    )
    ws.cell(row=headings_row, column=4).fill = fill(NAVY)
    ws.cell(row=headings_row, column=4).alignment = Alignment(
        horizontal="center", vertical="center"
    )
    ws.merge_cells(
        start_row=headings_row, start_column=4, end_row=headings_row, end_column=8
    )
    ws.cell(row=headings_row, column=9, value="NCC 2022 reference").font = title_font(
        11, color=WHITE
    )
    ws.cell(row=headings_row, column=9).fill = fill(NAVY)
    ws.cell(row=headings_row, column=9).alignment = Alignment(
        horizontal="center", vertical="center"
    )
    ws.row_dimensions[headings_row].height = 24

    for i, (sys_, desc, ref) in enumerate(items, start=1):
        r = headings_row + i
        a = ws.cell(row=r, column=2, value=sys_)
        a.font = body_font(10, bold=True)
        a.alignment = Alignment(wrap_text=True, vertical="center")
        a.border = ALL_THIN
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)

        b = ws.cell(row=r, column=4, value=desc)
        b.font = body_font(10)
        b.alignment = Alignment(wrap_text=True, vertical="center")
        b.border = ALL_THIN
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=8)

        c = ws.cell(row=r, column=9, value=ref)
        c.font = body_font(10, color=AMBER, bold=True)
        c.alignment = Alignment(wrap_text=True, vertical="center")
        c.border = ALL_THIN
        ws.row_dimensions[r].height = 38


# ---------------------------------------------------------------------------
# Sheet 4 — Type B scheme (3-storey Class 5 example)
# ---------------------------------------------------------------------------
def build_type_b(wb):
    ws = wb.create_sheet("Type B scheme")
    set_col_widths(ws, [3, 6, 8, 8, 8, 8, 18, 26, 26, 3])

    write_header(
        ws,
        "B2:I2",
        "Type B — Intermediate fire-resistance (NCC 2022 C2D2 + Spec 5 Table 4)",
        bg=ORANGE,
        size=14,
    )

    ws["B4"] = (
        "Design example: Class 2 residential building, 2 storeys above ground over a "
        "Class 7a garage, total height less than 12 m. Rise in Storeys = 2 → Type B "
        "(Class 2/3/9 column of Table C2D2). Garages and basements are excluded from "
        "the Rise in Storeys count under C2D4 when they meet the criteria of "
        "C2D4(2). Effective height < 12 m → does not trigger the >25 m fire safety "
        "package (no FCR, emergency lift, dual sprinkler supply)."
    )
    ws.merge_cells("B4:I4")
    ws["B4"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[4].height = 60

    levels = [
        ("L2", "Class 2 — Apartments (SOUs)", "Sole-occupancy unit walls 60/60/60 (Spec 5 Table 4)"),
        ("L1", "Class 2 — Apartments (SOUs)", "Apartment floor 60/60/60; corridor 60/60/60"),
        ("GF", "Lobby + Class 2 entry / mailboxes", "Two exits required from each storey above (D2D2)"),
        ("Garage", "Class 7a — Carpark (excluded from RIS)", "Sprinklered if >40 vehicles (Spec E1.5); separated from Class 2 by 90/90/90"),
    ]

    start_row = 6
    for i, (lbl, occ, note) in enumerate(levels):
        row = start_row + i
        style_facade_cell(ws.cell(row=row, column=2), NAVY)
        style_level_cell(ws.cell(row=row, column=3), lbl, CREAM)
        c = ws.cell(row=row, column=4, value=occ)
        c.fill = fill(LIGHT_GREY)
        c.font = body_font(9)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = ALL_THIN
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
        style_facade_cell(ws.cell(row=row, column=7), NAVY)
        n = ws.cell(row=row, column=8, value=note)
        n.font = body_font(9, color=AMBER)
        n.alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells(start_row=row, start_column=8, end_row=row, end_column=9)
        ws.row_dimensions[row].height = 28

    ground_row = start_row + len(levels)
    for col in range(2, 8):
        ws.cell(row=ground_row, column=col).fill = fill("8B7355")
        ws.cell(row=ground_row, column=col).border = ALL_THIN
    ws.row_dimensions[ground_row].height = 6

    sys_row = ground_row + 2
    write_header(
        ws,
        f"B{sys_row}:I{sys_row}",
        "Required fire safety systems — Type B, Class 2 less than 12 m",
        bg=AMBER,
        size=12,
    )

    items = [
        ("Fire isolated stair",
         "Required where the building has a Rise in Storeys of 2+ AND no other compliant exit is provided. Stair shaft FRL 90/90/90 (Spec 5 Table 4).",
         "D2D6, Spec 5 §3"),
        ("Hydrant system",
         "Required where total floor area > 500 m² OR effective height > 12 m. For typical sub-12 m Class 2 walk-up apartments, town main hydrants (within 90 m of FB access) are usually accepted.",
         "E1D2 / AS 2419.1"),
        ("Fire hose reels",
         "Required where a hydrant system is required, OR Class 2 building has total floor area > 500 m² and is more than 1 storey.",
         "E1D3 / AS 2441"),
        ("Sprinklers",
         "NOT required for a typical < 12 m Class 2 building. Becomes mandatory under E1D7 only when effective height > 25 m.",
         "E1D6, E1D7, Spec E1.5"),
        ("Smoke alarms",
         "Class 2 SOUs — interconnected mains-powered photoelectric smoke alarms (AS 3786) in every bedroom, hallway and storey.",
         "E2D2, Spec E2.2a"),
        ("SOU separation",
         "Walls between sole-occupancy units 60/60/60 with R w  + C tr ≥ 50; floors 60/60/60. Acoustic + fire double duty.",
         "C3D3, Spec 5 Table 4, F5D2"),
        ("External walls",
         "FRL 90/60/30 within 1.5 m of boundary; non-combustible cladding required for Class 2 buildings ≥ 3 storeys (C2D11).",
         "Spec 5 Table 4, C2D11"),
        ("Carpark separation",
         "Class 7a garage separated from the Class 2 part by walls and floor of 90/90/90 (Spec 5 §3.5).",
         "C3D8, Spec 5 §3.5"),
        ("Number of exits",
         "Two exits from each storey above ground (D2D2). Travel distance ≤ 6 m to a point of choice within an SOU; ≤ 20 m from point of choice to an exit.",
         "D2D2, D3D5"),
    ]

    headings_row = sys_row + 2
    ws.cell(row=headings_row, column=2, value="System / requirement").font = title_font(
        11, color=WHITE
    )
    ws.cell(row=headings_row, column=2).fill = fill(NAVY)
    ws.cell(row=headings_row, column=2).alignment = Alignment(
        horizontal="center", vertical="center"
    )
    ws.merge_cells(start_row=headings_row, start_column=2, end_row=headings_row, end_column=3)
    ws.cell(row=headings_row, column=4, value="Description").font = title_font(11, color=WHITE)
    ws.cell(row=headings_row, column=4).fill = fill(NAVY)
    ws.cell(row=headings_row, column=4).alignment = Alignment(
        horizontal="center", vertical="center"
    )
    ws.merge_cells(start_row=headings_row, start_column=4, end_row=headings_row, end_column=8)
    ws.cell(row=headings_row, column=9, value="NCC 2022 reference").font = title_font(
        11, color=WHITE
    )
    ws.cell(row=headings_row, column=9).fill = fill(NAVY)
    ws.cell(row=headings_row, column=9).alignment = Alignment(
        horizontal="center", vertical="center"
    )
    ws.row_dimensions[headings_row].height = 24

    for i, (sys_, desc, ref) in enumerate(items, start=1):
        r = headings_row + i
        a = ws.cell(row=r, column=2, value=sys_)
        a.font = body_font(10, bold=True)
        a.alignment = Alignment(wrap_text=True, vertical="center")
        a.border = ALL_THIN
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        b = ws.cell(row=r, column=4, value=desc)
        b.font = body_font(10)
        b.alignment = Alignment(wrap_text=True, vertical="center")
        b.border = ALL_THIN
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=8)
        c = ws.cell(row=r, column=9, value=ref)
        c.font = body_font(10, color=AMBER, bold=True)
        c.alignment = Alignment(wrap_text=True, vertical="center")
        c.border = ALL_THIN
        ws.row_dimensions[r].height = 38


# ---------------------------------------------------------------------------
# Sheet 5 — Type C scheme (single storey Class 5/6 example)
# ---------------------------------------------------------------------------
def build_type_c(wb):
    ws = wb.create_sheet("Type C scheme")
    set_col_widths(ws, [3, 6, 8, 8, 8, 8, 18, 26, 26, 3])

    write_header(
        ws,
        "B2:I2",
        "Type C — Least restrictive (NCC 2022 C2D2 + Spec 5 Table 5)",
        bg=GREEN,
        size=14,
    )

    ws["B4"] = (
        "Design example: Class 6 retail showroom, single storey, 950 m² floor area, "
        "with attached small Class 7b warehouse. Rise in Storeys = 1 → Type C. "
        "May use combustible structural framing (e.g. timber-framed) provided "
        "Spec 5 Table 5 FRLs and external wall fire source feature distances are met."
    )
    ws.merge_cells("B4:I4")
    ws["B4"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[4].height = 50

    levels = [
        ("Roof", "Lightweight roof — non-combustible if <900 mm to boundary", "C2D11"),
        ("L1", "Class 6 showroom + Class 7b storage", "Internal fire wall 90/90/90 between Class 6 and 7b"),
    ]

    start_row = 7
    for i, (lbl, occ, note) in enumerate(levels):
        row = start_row + i
        style_facade_cell(ws.cell(row=row, column=2), NAVY)
        style_level_cell(ws.cell(row=row, column=3), lbl, CREAM)
        c = ws.cell(row=row, column=4, value=occ)
        c.fill = fill(LIGHT_GREY)
        c.font = body_font(9)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = ALL_THIN
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
        style_facade_cell(ws.cell(row=row, column=7), NAVY)
        n = ws.cell(row=row, column=8, value=note)
        n.font = body_font(9, color=AMBER)
        n.alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells(start_row=row, start_column=8, end_row=row, end_column=9)
        ws.row_dimensions[row].height = 36

    ground_row = start_row + len(levels)
    for col in range(2, 8):
        ws.cell(row=ground_row, column=col).fill = fill("8B7355")
        ws.cell(row=ground_row, column=col).border = ALL_THIN
    ws.row_dimensions[ground_row].height = 6

    sys_row = ground_row + 2
    write_header(
        ws,
        f"B{sys_row}:I{sys_row}",
        "Required fire safety systems — Type C, single storey",
        bg=AMBER,
        size=12,
    )

    items = [
        ("Fire isolated stairs",
         "Generally not required (single storey). If a mezzanine: enclosed stair only when total area >200 m².",
         "D2D6"),
        ("Fire hydrant system",
         "Not required where total floor area ≤ 500 m². Showroom ~950 m² → may rely on town main hydrants ≤ 90 m from FB access.",
         "E1D2"),
        ("Fire hose reels",
         "Required where hydrants required, OR if Class 6 floor area >500 m².",
         "E1D3"),
        ("Sprinklers",
         "Not required unless triggered by Spec E1.5 (e.g. retail >2 000 m² in some classes, or large isolated buildings).",
         "E1D6, Spec E1.5"),
        ("Smoke alarms / detection",
         "Class 6 and 7b — heat alarms in tenancy; smoke alarms in any associated office/admin area.",
         "E2D2, Spec E2.2a"),
        ("FRL — loadbearing elements",
         "Generally 30/30/30 to internal columns and walls; 60/60/60 between adjoining tenancies of different classes.",
         "Spec 5 Table 5"),
        ("External walls",
         "FRL 60/60/60 within 1.5 m of boundary; non-combustible if <900 mm. May be combustible-framed beyond 3 m.",
         "Spec 5 Table 5, C2D11"),
        ("Number of exits",
         "Single exit permitted if travel distance ≤ 20 m and floor area ≤ 200 m² (D2D2). Otherwise 2 exits.",
         "D2D2, D3D5"),
        ("Compartmentation",
         "Class 6 maximum compartment 5 500 m²; Class 7b maximum 18 000 m³ volume.",
         "C3D3, Spec 5 Table 5"),
    ]

    headings_row = sys_row + 2
    ws.cell(row=headings_row, column=2, value="System / requirement").font = title_font(
        11, color=WHITE
    )
    ws.cell(row=headings_row, column=2).fill = fill(NAVY)
    ws.cell(row=headings_row, column=2).alignment = Alignment(
        horizontal="center", vertical="center"
    )
    ws.merge_cells(start_row=headings_row, start_column=2, end_row=headings_row, end_column=3)
    ws.cell(row=headings_row, column=4, value="Description").font = title_font(11, color=WHITE)
    ws.cell(row=headings_row, column=4).fill = fill(NAVY)
    ws.cell(row=headings_row, column=4).alignment = Alignment(
        horizontal="center", vertical="center"
    )
    ws.merge_cells(start_row=headings_row, start_column=4, end_row=headings_row, end_column=8)
    ws.cell(row=headings_row, column=9, value="NCC 2022 reference").font = title_font(
        11, color=WHITE
    )
    ws.cell(row=headings_row, column=9).fill = fill(NAVY)
    ws.cell(row=headings_row, column=9).alignment = Alignment(
        horizontal="center", vertical="center"
    )
    ws.row_dimensions[headings_row].height = 24

    for i, (sys_, desc, ref) in enumerate(items, start=1):
        r = headings_row + i
        a = ws.cell(row=r, column=2, value=sys_)
        a.font = body_font(10, bold=True)
        a.alignment = Alignment(wrap_text=True, vertical="center")
        a.border = ALL_THIN
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        b = ws.cell(row=r, column=4, value=desc)
        b.font = body_font(10)
        b.alignment = Alignment(wrap_text=True, vertical="center")
        b.border = ALL_THIN
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=8)
        c = ws.cell(row=r, column=9, value=ref)
        c.font = body_font(10, color=AMBER, bold=True)
        c.alignment = Alignment(wrap_text=True, vertical="center")
        c.border = ALL_THIN
        ws.row_dimensions[r].height = 38


# ---------------------------------------------------------------------------
# Sheet 6 — FRL summary (Spec 5 Tables 3, 4, 5)
# ---------------------------------------------------------------------------
def build_frl(wb):
    ws = wb.create_sheet("FRL summary")
    set_col_widths(ws, [3, 30, 18, 18, 18, 28, 3])

    write_header(
        ws,
        "B2:F2",
        "Specification 5 — Fire Resistance Levels by construction type",
        bg=NAVY,
        size=14,
    )
    ws["B4"] = (
        "FRL = Structural adequacy / Integrity / Insulation, in minutes. Values "
        "shown are the minimum DTS FRL for the worst-case Class served by that "
        "Type (typically Class 9b for Type A, Class 5/6 for Type B, Class 6/7 for "
        "Type C). Read with full Specification 5 tables for the exact Class."
    )
    ws.merge_cells("B4:F4")
    ws["B4"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[4].height = 40

    headers = ["Building element", "Type A (Spec 5 T3)", "Type B (T4)", "Type C (T5)", "Notes"]
    for i, h in enumerate(headers, start=2):
        c = ws.cell(row=6, column=i, value=h)
        c.fill = fill(GOLD)
        c.font = title_font(11, color=WHITE)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = ALL_THIN
    ws.row_dimensions[6].height = 32

    rows = [
        ("Loadbearing external wall (≤1.5 m to boundary)",
         "240/240/240", "120/120/120", "90/60/30",
         "Type A reaches 240 min for Class 9b assembly buildings."),
        ("Loadbearing external wall (1.5–3 m)",
         "240/180/120", "120/90/60", "60/30/30",
         "FRL reduces with distance from fire source feature."),
        ("Loadbearing external wall (>3 m)",
         "240/180/90", "90/60/30", "—/—/— (non-combustible only)",
         "Type C may use combustible framing beyond 3 m."),
        ("Internal loadbearing wall (incl. fire wall)",
         "240/240/240", "120/120/120", "60/60/60",
         "Fire wall must extend 900 mm above roof unless deemed-to-satisfy parapet."),
        ("Loadbearing column",
         "240/—/—", "120/—/—", "60/—/—",
         "Insulation/integrity often n/a for slender columns."),
        ("Floor",
         "120/120/120", "90/90/90", "60/60/60",
         "Minimum across all floor types in compartment."),
        ("Roof",
         "90/60/30", "60/30/30", "—/—/—",
         "Type C: roof FRL only if part of a fire-rated assembly."),
        ("Sole-occupancy unit separating wall (Class 2/3)",
         "90/90/90", "60/60/60", "60/60/60",
         "Acoustic and fire separation per F5 + Spec 5."),
    ]

    for r_idx, row in enumerate(rows, start=7):
        for c_idx, val in enumerate(row, start=2):
            c = ws.cell(row=r_idx, column=c_idx, value=val)
            c.alignment = Alignment(horizontal="center" if c_idx in (3, 4, 5) else "left",
                                    vertical="center", wrap_text=True)
            c.border = ALL_THIN
            if c_idx == 3:
                c.fill = fill("F8D7DA")  # light red
            elif c_idx == 4:
                c.fill = fill("FCE4CC")  # light orange
            elif c_idx == 5:
                c.fill = fill("D7E9CE")  # light green
            c.font = body_font(10)
        ws.row_dimensions[r_idx].height = 30


# ---------------------------------------------------------------------------
# Sheet 7 — Fire safety systems matrix
# ---------------------------------------------------------------------------
def build_systems_matrix(wb):
    ws = wb.create_sheet("Fire systems matrix")
    set_col_widths(ws, [3, 32, 14, 14, 14, 14, 28, 3])

    write_header(
        ws,
        "B2:G2",
        "Fire safety systems by effective height — extends original Sheet2",
        bg=NAVY,
        size=14,
    )

    headers = [
        "System / equipment",
        "<12 m (Type B/C)",
        "12 m – 25 m (Type A/B)",
        ">25 m (Type A)",
        ">50 m (Type A)",
        "Primary NCC reference",
    ]
    for i, h in enumerate(headers, start=2):
        c = ws.cell(row=4, column=i, value=h)
        c.fill = fill(GOLD)
        c.font = title_font(11, color=WHITE)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = ALL_THIN
    ws.row_dimensions[4].height = 40

    rows = [
        ("Fire hydrant system", "Optional (≤500 m²)", "Required", "Required", "Required", "E1D2 / AS 2419.1"),
        ("Fire hose reels", "Optional", "Required", "Required", "Required", "E1D3 / AS 2441"),
        ("Booster connection", "Optional", "Required", "Required", "Required (twin)", "E1D2 + AS 2419.1 §6"),
        ("Fire Control Room", "Not required", "Required (if FA >500 m²)", "Required", "Required (dedicated)", "E1D14 + Spec 20"),
        ("Fire isolated stairway / pressurised", "Not required", "Required", "Required (pressurised)", "Required (pressurised + open access ramp option)", "D2D6 / E2D14 / AS 1668.1"),
        ("Sprinkler system", "Per Spec E1.5", "Per Spec E1.5", "Required (Class 2/3)", "Required + dual water supply", "E1D6, E1D7"),
        ("Emergency lift", "Not required", "Optional", "Required", "Required (2 if >75 m)", "E3D9 / AS 1735.1"),
        ("Smoke hazard management", "Smoke alarms", "Smoke alarms + corridor detection", "Mechanical smoke control", "Smoke control + zone pressurisation", "E2D2 / Spec E2.2"),
        ("Operational hydrants during construction", "Not required", "Required when >12 m", "Required", "Required", "E1D2 + AS 2419.1"),
        ("Fire-protected timber (mass-timber)", "Permitted", "Permitted", "Restricted (≤25 m only)", "Not permitted in Class 2/3", "C2D14 + Spec 7"),
    ]

    for r_idx, row in enumerate(rows, start=5):
        for c_idx, val in enumerate(row, start=2):
            c = ws.cell(row=r_idx, column=c_idx, value=val)
            c.alignment = Alignment(horizontal="left" if c_idx in (2, 7) else "center",
                                    vertical="center", wrap_text=True)
            c.border = ALL_THIN
            c.font = body_font(10)
            if c_idx in (3, 4, 5, 6):
                v = (val or "").lower()
                if "required" in v and "not required" not in v:
                    c.fill = fill("D7E9CE")
                elif "optional" in v or "permitted" in v:
                    c.fill = fill("FFF6CC")
                elif "not required" in v or "not permitted" in v or "restricted" in v:
                    c.fill = fill("F8D7DA")
        ws.row_dimensions[r_idx].height = 32


# ---------------------------------------------------------------------------
# Sheet 8 — Compartmentation & exits
# ---------------------------------------------------------------------------
def build_compartments(wb):
    ws = wb.create_sheet("Compartments & exits")
    set_col_widths(ws, [3, 18, 26, 18, 18, 22, 3])

    write_header(
        ws,
        "B2:F2",
        "Compartmentation, travel distances & exits (NCC 2022 C3 + D2/D3)",
        bg=NAVY,
        size=14,
    )

    headers = ["Class", "Description", "Max compartment area (m²)", "Max travel to point of choice (m)", "Min exits per storey"]
    for i, h in enumerate(headers, start=2):
        c = ws.cell(row=4, column=i, value=h)
        c.fill = fill(GOLD)
        c.font = title_font(11, color=WHITE)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = ALL_THIN
    ws.row_dimensions[4].height = 36

    rows = [
        ("Class 2", "Sole-occupancy units (apartments)", "n/a — SOU separation 90/90/90", "6 m to a SOU; 20 m to exit", "2 (above 2 storeys)"),
        ("Class 3", "Hotel, hostel, residential care", "8 000 m²", "20 m", "2"),
        ("Class 5", "Office", "8 000 m²", "20 m", "2"),
        ("Class 6", "Shop / retail", "5 500 m²", "20 m", "2"),
        ("Class 7a", "Carpark", "8 000 m² (sprinklered)", "40 m (sprinklered)", "2"),
        ("Class 7b", "Storage", "18 000 m³ volume", "40 m", "2"),
        ("Class 8", "Factory / lab", "8 000 m²", "40 m", "2"),
        ("Class 9a", "Hospital / health-care", "2 000 m²", "20 m (12 m in patient care)", "2"),
        ("Class 9b", "Assembly", "8 000 m²", "20 m", "2 (1 if ≤25 occupants in single storey)"),
        ("Class 9c", "Aged care", "500 m² (resident-use area)", "20 m (10 m within smoke compartment)", "2"),
    ]

    for r_idx, row in enumerate(rows, start=5):
        for c_idx, val in enumerate(row, start=2):
            c = ws.cell(row=r_idx, column=c_idx, value=val)
            c.alignment = Alignment(horizontal="center" if c_idx > 3 else "left",
                                    vertical="center", wrap_text=True)
            c.border = ALL_THIN
            c.font = body_font(10)
            if c_idx == 2:
                c.fill = fill(SKY)
        ws.row_dimensions[r_idx].height = 30

    ws["B16"] = "Notes"
    ws["B16"].font = body_font(11, color=AMBER, bold=True)
    notes = (
        "• Travel distances and exit numbers are minimums. Where a Performance Solution "
        "is used, equivalent or better outcomes must be demonstrated under A2G2.\n"
        "• Compartment areas reduce when sprinklers are not provided; the values shown "
        "assume the building is sprinkler-protected where E1D6 / Spec E1.5 require it.\n"
        "• Class 9c values reflect the special concession in C2D2(c) and Spec 9.\n"
        "• Atrium, large isolated buildings and open-deck carparks have additional "
        "concessions in Spec C3.4, Spec G3, Spec G6."
    )
    ws.merge_cells("B17:F22")
    ws["B17"] = notes
    ws["B17"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["B17"].font = body_font(10)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main(out_path):
    wb = Workbook()
    build_cover(wb)
    build_c2d2(wb)
    build_type_a(wb)
    build_type_b(wb)
    build_type_c(wb)
    build_frl(wb)
    build_systems_matrix(wb)
    build_compartments(wb)
    wb.save(out_path)
    print(f"Wrote: {out_path}")


if __name__ == "__main__":
    import sys
    out = sys.argv[1] if len(sys.argv) > 1 else "Building_Types_NCC2022.xlsx"
    main(out)
