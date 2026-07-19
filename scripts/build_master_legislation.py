"""
369 Alliance — Master NCC + AS Legislation Reference Builder
============================================================

Combines four authoritative sources into a single, fully cross-referenced
Excel workbook used by inspectors, auditors, designers and certifiers:

  1. NCC Building Reference Guide xlsx       (977-row 2019 cross-reference)
  2. AS in use.docx                          (38 categorised AS standards)
  3. NCC 2022 Volume One PDF + 82 clause CSVs (Sections B-J, Spec 1-45)
  4. NCC 2022 Volume Two (Housing Provisions) clause list (337 clauses)
  5. NCC 2022 Volume Two referenced standards (237 entries)
  6. Master Defect Library (182 defect items)

Output: C:\\2026\\369 Alliance\\1. Working folder\\
          MASTER_NCC_AS_Legislation_Reference.xlsx
"""
from __future__ import annotations
import csv
import re
import sys
from pathlib import Path
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8')

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule, ColorScaleRule
from openpyxl.worksheet.table import Table, TableStyleInfo
import docx

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
ROOT = Path(r"C:\2026\369 Alliance")
LEG  = ROOT / "1. Legislation"
NCC  = LEG / "NCC"
WORK = ROOT / "1. Working folder"
OUT  = WORK / "MASTER_NCC_AS_Legislation_Reference.xlsx"

# ---------------------------------------------------------------------------
# Brand styling — 369 Alliance palette
# ---------------------------------------------------------------------------
NAVY      = "1A1A2E"
GOLD      = "A68A64"
AMBER     = "C07040"
CREAM     = "F8F4ED"
LIGHTGOLD = "EFE3CC"
WHITE     = "FFFFFF"
GREY_HDR  = "2C2C40"
GREY_LITE = "F2F2F2"

# Section colour scheme (Volume One)
SECTION_COLOURS = {
    "A": "5B6C8C",  # General provisions — slate
    "B": "8B5A3C",  # Structure — earthy brown
    "C": "B22222",  # Fire resistance — fire red
    "D": "FF8C00",  # Access & egress — orange
    "E": "DAA520",  # Services & equipment — goldenrod
    "F": "1E90FF",  # Health & amenity — blue
    "G": "228B22",  # Ancillary provisions — green
    "I": "9370DB",  # Special use buildings — purple
    "J": "20B2AA",  # Energy efficiency — teal
    "S": "696969",  # Specifications — dim grey
    "V2": "4F6228", # Housing Provisions / Vol 2
}

REGIME_COLOURS = {
    "Waterproofing":          "1E90FF",  # blue
    "Fire safety":            "B22222",  # red
    "Fire":                   "B22222",
    "Structure":              "8B5A3C",
    "Internal or external load-bearing component": "8B5A3C",
    "Building enclosure":     "DAA520",
    "Mechanical, plumbing and electrical service": "20B2AA",
    "Essential Services":     "20B2AA",
    "Cladding":               "C07040",
}

# ---------------------------------------------------------------------------
# Common styles
# ---------------------------------------------------------------------------
THIN = Side(style="thin", color="BFBFBF")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def hdr_fill(hex_): return PatternFill("solid", fgColor=hex_)
def font(hex_=None, bold=False, size=11, italic=False, name="Calibri"):
    kw = dict(bold=bold, size=size, italic=italic, name=name)
    if hex_: kw["color"] = hex_
    return Font(**kw)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
LEFTTOP= Alignment(horizontal="left",   vertical="top",    wrap_text=True)

def write_header_row(ws, row_idx, headers, fill_hex=NAVY, font_hex=WHITE):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row_idx, column=c, value=h)
        cell.fill = hdr_fill(fill_hex)
        cell.font = font(font_hex, bold=True, size=11)
        cell.alignment = CENTER
        cell.border = BORDER_ALL
    ws.row_dimensions[row_idx].height = 32

def autosize(ws, max_widths=None):
    """Set column widths from observed content (capped)."""
    max_widths = max_widths or {}
    for col_idx, col in enumerate(ws.columns, 1):
        letter = get_column_letter(col_idx)
        max_len = 10
        for cell in col:
            try:
                v = str(cell.value) if cell.value is not None else ""
            except Exception:
                v = ""
            for line in v.split("\n"):
                if len(line) > max_len:
                    max_len = len(line)
        cap = max_widths.get(letter, 60)
        ws.column_dimensions[letter].width = min(max_len + 2, cap)

def add_filter_and_freeze(ws, header_row=1, freeze_at=None):
    last_col = get_column_letter(ws.max_column)
    ws.auto_filter.ref = f"A{header_row}:{last_col}{ws.max_row}"
    ws.freeze_panes = freeze_at or f"A{header_row+1}"

def stripe_rows(ws, start_row=2, fill_hex=GREY_LITE):
    for r in range(start_row, ws.max_row + 1):
        if (r - start_row) % 2 == 1:
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                if cell.fill.fgColor.rgb in (None, "00000000", "FFFFFFFF"):
                    cell.fill = hdr_fill(fill_hex)

# ---------------------------------------------------------------------------
# Source loaders
# ---------------------------------------------------------------------------
def load_csv_labels(path: Path) -> list[str]:
    """Read a single-column 'label' csv from the NCC dataset."""
    out = []
    with open(path, encoding="utf-8") as f:
        for i, row in enumerate(csv.reader(f)):
            if i == 0 and row and row[0].lstrip("﻿").strip().lower() == "label":
                continue
            if not row: continue
            text = " ".join([c for c in row if c]).strip()
            text = text.lstrip("﻿").strip()
            if text:
                out.append(text)
    return out

# Volume One: parse "B1D2 Resistance to actions" or "S26C3 Shower area..."
CLAUSE_VOL1 = re.compile(r"^(?P<code>[A-Z]+\d*[A-Z]?\d*)\s+(?P<title>.+)$")

def parse_vol1_filename(fname: str) -> tuple[str, str, str, str]:
    """Return (section_letter, part_or_spec, part_title, kind)."""
    base = fname.replace(".csv", "")
    # Specifications
    m = re.search(r"Specification_(\d+)_(.+)$", base)
    if m:
        num = m.group(1)
        title = m.group(2).replace("_", " ")
        return ("S", f"S{num}", f"Specification {num} — {title}", "Specification")
    # Parts: NCC_2022_Clauses_001_Part_B1_Structural_provisions
    m = re.search(r"Part_([A-Z]\d?)_?(.*)$", base)
    if m:
        part_code = m.group(1)
        title = m.group(2).replace("_", " ").strip() or part_code
        section = part_code[0]
        return (section, part_code, f"Part {part_code} — {title}", "Part")
    return ("?", "?", base, "?")

CLAUSE_TYPE_MAP = {
    "O": "Objective",
    "F": "Functional Statement",
    "P": "Performance Requirement",
    "V": "Verification Method",
    "D": "Deemed-to-Satisfy",
    "G": "Governing / General",
    "C": "Specification Clause",
}

def classify_clause_code(code: str) -> str:
    """Detect clause type from BxDy / SxCy / xPy patterns."""
    # Vol 1 clauses: section letter + part digit + type letter + number, e.g. B1D2, C2P1
    m = re.match(r"^[A-Z]+\d+([A-Z])\d+", code)
    if m:
        return CLAUSE_TYPE_MAP.get(m.group(1), "Provision")
    # Specifications S26C3 -> 'C' = clause
    m = re.match(r"^S\d+([A-Z])\d+", code)
    if m:
        return CLAUSE_TYPE_MAP.get(m.group(1), "Specification Clause")
    return "Provision"

# ---------------------------------------------------------------------------
# Build NCC Volume 1 dataset
# ---------------------------------------------------------------------------
def build_ncc_vol1():
    rows = []
    files = sorted(NCC.glob("NCC_2022_Clauses_*.csv"))
    for f in files:
        section, part_code, part_title, kind = parse_vol1_filename(f.name)
        labels = load_csv_labels(f)
        for raw in labels:
            # Try splitting "CODE Title"
            m = CLAUSE_VOL1.match(raw)
            if m:
                code = m.group("code")
                title = m.group("title")
            else:
                code = ""
                title = raw
            ctype = classify_clause_code(code) if code else ""
            rows.append({
                "ncc_version": "NCC 2022",
                "volume":      "Volume One",
                "section":     section,
                "part_code":   part_code,
                "part_title":  part_title,
                "kind":        kind,
                "clause_code": code,
                "clause_type": ctype,
                "clause_title": title,
                "source_file": f.name,
            })
    return rows

# ---------------------------------------------------------------------------
# Build NCC Volume 2 (Housing Provisions) dataset
# ---------------------------------------------------------------------------
HP_PARSE = re.compile(
    r"^(?P<sec>\d+)\s+(?P<sec_title>[^-–>]+?)\s*[-–>]+\s*"
    r"(?:Part\s+(?P<part>\d+(?:\.\d+)?)\s*(?P<part_title>[^-–>]*?))?"
    r"\s*(?:[-–>]+\s*(?P<clause>\d+(?:\.\d+)+)\s*(?P<clause_title>.+))?$"
)

def build_ncc_vol2():
    rows = []
    p = NCC / "NCC_2022_HP_allClause.csv"
    labels = load_csv_labels(p)
    for raw in labels:
        # Replace ' > ' with ' - ' for unified parsing
        clean = raw.replace(" > ", " - ").replace("  ", " ").strip()
        m = HP_PARSE.match(clean)
        if m:
            rows.append({
                "ncc_version": "NCC 2022",
                "volume":      "Volume Two (Housing Provisions)",
                "section":     m.group("sec") or "",
                "section_title": (m.group("sec_title") or "").strip(),
                "part_code":   (m.group("part") or "").strip(),
                "part_title":  (m.group("part_title") or "").strip(),
                "clause_code": (m.group("clause") or "").strip(),
                "clause_title": (m.group("clause_title") or "").strip() or (m.group("part_title") or "").strip() or (m.group("sec_title") or "").strip(),
                "source": "NCC_2022_HP_allClause.csv",
            })
        else:
            rows.append({
                "ncc_version": "NCC 2022",
                "volume":      "Volume Two (Housing Provisions)",
                "section": "", "section_title": "",
                "part_code": "", "part_title": "",
                "clause_code": "",
                "clause_title": clean,
                "source": "NCC_2022_HP_allClause.csv",
            })
    return rows

# ---------------------------------------------------------------------------
# AS Standards (Vol 2 referenced standards + AS in use)
# ---------------------------------------------------------------------------
STD_PARSE = re.compile(r"^\[(?P<vstr>NCC\d{4})\]\s*(?P<rest>.+)$")

def build_referenced_standards():
    rows = []
    p = NCC / "NCC Volume 2 Standards.csv"
    with open(p, encoding="utf-8") as f:
        for i, row in enumerate(csv.reader(f)):
            if i == 0: continue
            if not row: continue
            text = " ".join([c.strip() for c in row if c]).strip()
            text = text.lstrip("﻿").strip()
            if not text: continue
            m = STD_PARSE.match(text)
            if not m:
                rows.append({
                    "ncc_version": "Unknown", "code": "", "year": "",
                    "title": text, "version_note": "", "source": "NCC Volume 2 Standards.csv"
                })
                continue
            ver = m.group("vstr")  # NCC2019 or NCC2022
            rest = m.group("rest")
            # rest is like "AS 1170 Part 4 |  2024 |  Title |  Version note"
            parts = [p.strip() for p in rest.split("|")]
            code = parts[0] if parts else ""
            year = parts[1] if len(parts) > 1 else ""
            title = parts[2] if len(parts) > 2 else ""
            note  = parts[3] if len(parts) > 3 else ""
            rows.append({
                "ncc_version": ver, "code": code, "year": year,
                "title": title, "version_note": note,
                "source": "NCC Volume 2 Standards.csv",
            })
    return rows

def build_as_in_use():
    """AS in use.docx — categorised list."""
    rows = []
    d = docx.Document(LEG / "AS in use.docx")
    t = d.tables[0]
    current_cat = ""
    for r in t.rows:
        a = r.cells[0].text.strip()
        b = r.cells[1].text.strip()
        if a == b and a:
            current_cat = a
        elif a:
            rows.append({"category": current_cat, "code": a, "title": b})
    return rows

# ---------------------------------------------------------------------------
# 2019 Reference Guide -> Topic / NCC / AS crosswalk
# ---------------------------------------------------------------------------
def build_crosswalk():
    rows = []
    wb = openpyxl.load_workbook(LEG / "NCC Building Reference Guide-Updated with new links 21 October 2021.xlsx", data_only=True)
    if "2019" not in wb.sheetnames:
        wb.close()
        return rows
    ws = wb["2019"]
    current_topic = ""
    for ridx in range(2, ws.max_row + 1):
        item = ws.cell(ridx, 1).value
        if item is None:
            continue
        if isinstance(item, str) and item.strip().startswith("*"):
            current_topic = item.strip().lstrip("*").strip()
            continue
        if not isinstance(item, str) or not item.strip():
            continue
        rows.append({
            "topic": current_topic,
            "complaint_item": item.strip(),
            "ncc_vol1_2019":  (ws.cell(ridx, 2).value or "").strip() if ws.cell(ridx, 2).value else "",
            "ncc_vol2_2019":  (ws.cell(ridx, 3).value or "").strip() if ws.cell(ridx, 3).value else "",
            "ncc_vol3_2019":  (ws.cell(ridx, 4).value or "").strip() if ws.cell(ridx, 4).value else "",
            "as_reference":   (ws.cell(ridx, 5).value or "").strip() if ws.cell(ridx, 5).value else "",
            "guide_to_std":   (ws.cell(ridx, 6).value or "").strip() if ws.cell(ridx, 6).value else "",
            "other":          (ws.cell(ridx, 7).value or "").strip() if ws.cell(ridx, 7).value else "",
            "previous":       (ws.cell(ridx, 8).value or "").strip() if ws.cell(ridx, 8).value else "",
        })
    wb.close()
    return rows

# ---------------------------------------------------------------------------
# Defect Library
# ---------------------------------------------------------------------------
def build_defects():
    rows = []
    wb = openpyxl.load_workbook(WORK / "New Master_Defect Library iAuditor_r.xlsx", data_only=True)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    for r in range(2, ws.max_row + 1):
        rec = {}
        for c, h in enumerate(headers, 1):
            v = ws.cell(r, c).value
            if isinstance(v, str): v = v.strip()
            rec[h] = v
        if rec.get("Defect ID"):
            rows.append(rec)
    wb.close()
    return rows

# ---------------------------------------------------------------------------
# Workbook builder
# ---------------------------------------------------------------------------
def main():
    print(f"[ ] Loading sources …")
    vol1     = build_ncc_vol1()
    print(f"  · Vol 1 clauses           : {len(vol1)}")
    vol2     = build_ncc_vol2()
    print(f"  · Vol 2 (HP) clauses      : {len(vol2)}")
    refstd   = build_referenced_standards()
    print(f"  · Vol 2 referenced std    : {len(refstd)}")
    asuse    = build_as_in_use()
    print(f"  · AS in use (categorised) : {len(asuse)}")
    cross    = build_crosswalk()
    print(f"  · Topic-NCC-AS crosswalk  : {len(cross)}")
    defects  = build_defects()
    print(f"  · Defect library entries  : {len(defects)}")

    wb = Workbook()
    wb.remove(wb.active)

    # =====================================================================
    # SHEET 1 — Cover
    # =====================================================================
    ws = wb.create_sheet("Cover", 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 38
    ws.column_dimensions['C'].width = 78
    ws.column_dimensions['D'].width = 18

    ws.merge_cells("B2:D2")
    c = ws["B2"]
    c.value = "369 ALLIANCE — Master NCC + Australian Standards Reference"
    c.font = font(NAVY, bold=True, size=22)
    c.alignment = LEFT
    ws.row_dimensions[2].height = 32

    ws.merge_cells("B3:D3")
    c = ws["B3"]
    c.value = "The definitive cross-referenced legislation register for NSW Class 2/3/9c construction compliance."
    c.font = font(GOLD, bold=False, size=12, italic=True)
    c.alignment = LEFT

    ws.merge_cells("B4:D4")
    c = ws["B4"]
    c.value = f"Compiled {datetime.now().strftime('%d %b %Y')}  ·  369 Alliance, Sydney NSW"
    c.font = font(GREY_HDR, size=10)
    c.alignment = LEFT

    # Section break
    ws.merge_cells("B6:D6")
    c = ws["B6"]; c.value = "WORKBOOK CONTENTS"
    c.font = font(WHITE, bold=True, size=12); c.fill = hdr_fill(NAVY); c.alignment = LEFT
    c.border = BORDER_ALL

    sheets_meta = [
        ("01 NCC 2022 Vol 1 Clauses",      f"All {len(vol1)} clauses across Sections B-J + 45 Specifications"),
        ("02 NCC 2022 Vol 2 (HP)",         f"All {len(vol2)} Housing Provisions clauses (Sections 2-13)"),
        ("03 AS Standards Master",          f"All {len(refstd)} NCC-referenced standards (NCC2019 + NCC2022)"),
        ("04 AS in Use (categorised)",     f"{len(asuse)} active AS standards organised by discipline"),
        ("05 Topic-NCC-AS Crosswalk",      f"{len(cross)} complaint-driven cross-references (BCA Vol 1/2/3 + AS)"),
        ("06 Master Defect Library",       f"{len(defects)} regulated defect items with NCC 2019 + 2022 mapping"),
        ("07 Quick Lookup by Element",     "Element/regime → all relevant NCC clauses + AS standards"),
        ("08 Reference Standards Index",   "All AS / referenced documents indexed by code"),
        ("09 Section Colour Legend",       "Visual key for NCC sections / regimes / clause types"),
    ]
    r = 7
    for s_name, s_desc in sheets_meta:
        ws.cell(r, 2, s_name).font = font(NAVY, bold=True)
        ws.cell(r, 2).alignment = LEFT
        ws.cell(r, 3, s_desc).font = font(GREY_HDR)
        ws.cell(r, 3).alignment = LEFT
        # Hyperlink will be added later (after sheets exist) — do it inline below by storing label
        r += 1

    # Sources block
    r += 1
    ws.merge_cells(f"B{r}:D{r}"); c = ws.cell(r, 2)
    c.value = "AUTHORITATIVE SOURCES MERGED"
    c.font = font(WHITE, bold=True, size=12); c.fill = hdr_fill(NAVY); c.alignment = LEFT
    c.border = BORDER_ALL
    r += 1
    sources = [
        ("NCC 2022 Volume One (PDF)",     r"1. Legislation\NCC\PDF\NCC-2022-Volume-One-PCD.pdf"),
        ("NCC 2022 Vol 1 Clause datasets", "82 CSV exports — Sections B-J + Specifications 1-45"),
        ("NCC 2022 Vol 2 HP all clauses",  r"1. Legislation\NCC\NCC_2022_HP_allClause.csv"),
        ("NCC Vol 2 referenced standards", r"1. Legislation\NCC\NCC Volume 2 Standards.csv"),
        ("NCC Building Reference Guide",   r"1. Legislation\NCC Building Reference Guide-Updated with new links 21 October 2021.xlsx"),
        ("Australian Standards in use",    r"1. Legislation\AS in use.docx"),
        ("Master Defect Library iAuditor", r"1. Working folder\New Master_Defect Library iAuditor_r.xlsx"),
    ]
    for label, path in sources:
        ws.cell(r, 2, label).font = font(NAVY, bold=True)
        ws.cell(r, 3, path).font = font(GREY_HDR, italic=True, size=10)
        r += 1

    # Statistics block
    r += 1
    ws.merge_cells(f"B{r}:D{r}"); c = ws.cell(r, 2)
    c.value = "REGISTER STATISTICS"
    c.font = font(WHITE, bold=True, size=12); c.fill = hdr_fill(NAVY); c.alignment = LEFT
    c.border = BORDER_ALL
    r += 1
    stats = [
        ("Total NCC 2022 Vol 1 clauses",  len(vol1)),
        ("Total NCC 2022 Vol 2 clauses",  len(vol2)),
        ("Total referenced standards",    len(refstd)),
        ("Total categorised AS in use",   len(asuse)),
        ("Total topic crosswalk entries", len(cross)),
        ("Total defect library items",    len(defects)),
        ("GRAND TOTAL legislation rows", len(vol1)+len(vol2)+len(refstd)+len(asuse)+len(cross)+len(defects)),
    ]
    for label, n in stats:
        ws.cell(r, 2, label).font = font(NAVY)
        ws.cell(r, 3, n).font = font(AMBER, bold=True, size=12)
        ws.cell(r, 3).alignment = LEFT
        r += 1

    # Methodology block
    r += 1
    ws.merge_cells(f"B{r}:D{r}"); c = ws.cell(r, 2)
    c.value = "METHODOLOGY & USE"
    c.font = font(WHITE, bold=True, size=12); c.fill = hdr_fill(NAVY); c.alignment = LEFT
    c.border = BORDER_ALL
    r += 1
    methodology = [
        "Every clause carries the originating Section / Part code so it can be traced back to the NCC.",
        "Performance Requirements (P), Verification Methods (V) and Deemed-to-Satisfy (D) provisions are colour-coded.",
        "AS standards are duplicated where they appear in both NCC 2019 and NCC 2022 — version of currency is preserved.",
        "Defect codes follow the NSW Building Commissioner regimes: W (Waterproofing), F (Fire), S (Structural), BE (Building Enclosure), ES (Essential Services).",
        "All sheets have auto-filter and frozen header rows. Use Data → Filter to drill into any field.",
        "Use sheet 07 'Quick Lookup' to start from a building element and find every applicable code.",
    ]
    for m in methodology:
        ws.merge_cells(f"B{r}:D{r}")
        ws.cell(r, 2, "•  " + m).font = font(GREY_HDR, size=10)
        ws.cell(r, 2).alignment = LEFTTOP
        ws.row_dimensions[r].height = 22
        r += 1

    # =====================================================================
    # SHEET 02 — NCC 2022 Vol 1 Clauses (build before so we can link from cover)
    # =====================================================================
    s_vol1 = wb.create_sheet("01 NCC 2022 Vol 1 Clauses")
    headers = ["NCC Version", "Volume", "Section", "Part Code", "Part Title",
               "Kind", "Clause Code", "Clause Type", "Clause Title", "Source File"]
    write_header_row(s_vol1, 1, headers, NAVY, WHITE)
    for i, row in enumerate(vol1, 2):
        s_vol1.cell(i, 1, row["ncc_version"])
        s_vol1.cell(i, 2, row["volume"])
        sc = s_vol1.cell(i, 3, row["section"])
        sc.fill = hdr_fill(SECTION_COLOURS.get(row["section"], "BFBFBF"))
        sc.font = font(WHITE, bold=True); sc.alignment = CENTER
        s_vol1.cell(i, 4, row["part_code"]).font = font(NAVY, bold=True)
        s_vol1.cell(i, 5, row["part_title"])
        s_vol1.cell(i, 6, row["kind"])
        cc = s_vol1.cell(i, 7, row["clause_code"]); cc.font = font(NAVY, bold=True)
        ct = s_vol1.cell(i, 8, row["clause_type"])
        # Colour by clause type
        type_colour = {
            "Objective":            "B0C4DE",
            "Functional Statement": "C2E0FF",
            "Performance Requirement": "F4B084",
            "Verification Method":  "FFD966",
            "Deemed-to-Satisfy":    "C6EFCE",
            "Specification Clause": "D9D9D9",
            "Governing / General":  "E7E6E6",
        }.get(row["clause_type"], None)
        if type_colour:
            ct.fill = hdr_fill(type_colour); ct.alignment = CENTER
        s_vol1.cell(i, 9, row["clause_title"]).alignment = LEFTTOP
        s_vol1.cell(i,10, row["source_file"]).font = font(GREY_HDR, size=9, italic=True)
        for col in range(1, len(headers)+1):
            s_vol1.cell(i, col).border = BORDER_ALL
            if col not in (3, 8):
                s_vol1.cell(i, col).alignment = LEFTTOP
    autosize(s_vol1, max_widths={"E": 50, "I": 80, "J": 35, "B": 14})
    add_filter_and_freeze(s_vol1)

    # =====================================================================
    # SHEET 03 — NCC 2022 Vol 2 (Housing Provisions)
    # =====================================================================
    s_vol2 = wb.create_sheet("02 NCC 2022 Vol 2 (HP)")
    headers = ["NCC Version", "Volume", "Section #", "Section Title",
               "Part Code", "Part Title", "Clause Code", "Clause Title", "Source"]
    write_header_row(s_vol2, 1, headers, "4F6228", WHITE)
    for i, row in enumerate(vol2, 2):
        s_vol2.cell(i, 1, row["ncc_version"])
        s_vol2.cell(i, 2, row["volume"])
        sc = s_vol2.cell(i, 3, row["section"]); sc.alignment = CENTER
        sc.fill = hdr_fill(SECTION_COLOURS["V2"]); sc.font = font(WHITE, bold=True)
        s_vol2.cell(i, 4, row["section_title"]).font = font(NAVY, bold=True)
        s_vol2.cell(i, 5, row["part_code"]).alignment = CENTER
        s_vol2.cell(i, 6, row["part_title"])
        s_vol2.cell(i, 7, row["clause_code"]).font = font(NAVY, bold=True)
        s_vol2.cell(i, 7).alignment = CENTER
        s_vol2.cell(i, 8, row["clause_title"])
        s_vol2.cell(i, 9, row["source"]).font = font(GREY_HDR, size=9, italic=True)
        for col in range(1, len(headers)+1):
            s_vol2.cell(i, col).border = BORDER_ALL
            if col not in (3, 5, 7):
                s_vol2.cell(i, col).alignment = LEFTTOP
    autosize(s_vol2, max_widths={"D": 36, "F": 50, "H": 80, "I": 32, "B": 30})
    add_filter_and_freeze(s_vol2)

    # =====================================================================
    # SHEET 04 — AS Standards Master (NCC referenced)
    # =====================================================================
    s_std = wb.create_sheet("03 AS Standards Master")
    headers = ["NCC Version", "Standard Code", "Year", "Title", "Version / Notes", "Source"]
    write_header_row(s_std, 1, headers, "8B5A3C", WHITE)
    for i, row in enumerate(refstd, 2):
        v = row["ncc_version"]
        v_cell = s_std.cell(i, 1, v); v_cell.alignment = CENTER
        if v == "NCC2022":
            v_cell.fill = hdr_fill("C6EFCE"); v_cell.font = font(NAVY, bold=True)
        elif v == "NCC2019":
            v_cell.fill = hdr_fill("FFE699"); v_cell.font = font(NAVY, bold=True)
        s_std.cell(i, 2, row["code"]).font = font(NAVY, bold=True)
        s_std.cell(i, 3, row["year"]).alignment = CENTER
        s_std.cell(i, 4, row["title"]).alignment = LEFTTOP
        s_std.cell(i, 5, row["version_note"]).font = font(GREY_HDR, italic=True)
        s_std.cell(i, 6, row["source"]).font = font(GREY_HDR, size=9)
        for col in range(1, len(headers)+1):
            s_std.cell(i, col).border = BORDER_ALL
    autosize(s_std, max_widths={"D": 70, "E": 30, "F": 32, "B": 32})
    add_filter_and_freeze(s_std)

    # =====================================================================
    # SHEET 05 — AS in Use (categorised)
    # =====================================================================
    s_asu = wb.create_sheet("04 AS in Use (categorised)")
    headers = ["Discipline / Category", "AS Code", "Title"]
    write_header_row(s_asu, 1, headers, GOLD, WHITE)
    cat_colours = {
        "Waterproofing": "1E90FF", "Fire": "B22222", "Structure": "8B5A3C",
        "Wind Classification": "20B2AA", "Cladding": "C07040",
        "Essential Services": "DAA520",
    }
    for i, row in enumerate(asuse, 2):
        cat = row["category"]
        c = s_asu.cell(i, 1, cat); c.alignment = CENTER
        if cat in cat_colours:
            c.fill = hdr_fill(cat_colours[cat]); c.font = font(WHITE, bold=True)
        s_asu.cell(i, 2, row["code"]).font = font(NAVY, bold=True)
        s_asu.cell(i, 3, row["title"]).alignment = LEFTTOP
        for col in range(1, len(headers)+1):
            s_asu.cell(i, col).border = BORDER_ALL
    autosize(s_asu, max_widths={"A": 28, "B": 26, "C": 90})
    add_filter_and_freeze(s_asu)

    # =====================================================================
    # SHEET 06 — Topic-NCC-AS Crosswalk
    # =====================================================================
    s_cw = wb.create_sheet("05 Topic-NCC-AS Crosswalk")
    headers = ["Topic / Family", "Complaint Item / Sub-issue",
               "BCA Vol 1 (2019)", "BCA Vol 2 (2019)", "BCA Vol 3 (2019)",
               "Australian Standard", "Guide to Standards & Tolerances", "Other", "Previous Versions"]
    write_header_row(s_cw, 1, headers, AMBER, WHITE)
    for i, row in enumerate(cross, 2):
        s_cw.cell(i, 1, row["topic"]).font = font(NAVY, bold=True)
        s_cw.cell(i, 2, row["complaint_item"])
        s_cw.cell(i, 3, row["ncc_vol1_2019"])
        s_cw.cell(i, 4, row["ncc_vol2_2019"])
        s_cw.cell(i, 5, row["ncc_vol3_2019"])
        s_cw.cell(i, 6, row["as_reference"])
        s_cw.cell(i, 7, row["guide_to_std"])
        s_cw.cell(i, 8, row["other"])
        s_cw.cell(i, 9, row["previous"]).font = font(GREY_HDR, italic=True, size=9)
        for col in range(1, len(headers)+1):
            s_cw.cell(i, col).border = BORDER_ALL
            s_cw.cell(i, col).alignment = LEFTTOP
    autosize(s_cw, max_widths={"A": 26, "B": 50, "C": 45, "D": 45, "E": 30, "F": 45, "G": 35, "H": 25, "I": 30})
    add_filter_and_freeze(s_cw)

    # =====================================================================
    # SHEET 07 — Master Defect Library
    # =====================================================================
    s_def = wb.create_sheet("06 Master Defect Library")
    if defects:
        headers = list(defects[0].keys())
        write_header_row(s_def, 1, headers, "8B5A3C", WHITE)
        for i, rec in enumerate(defects, 2):
            for c, h in enumerate(headers, 1):
                v = rec.get(h, "")
                cell = s_def.cell(i, c, v)
                cell.alignment = LEFTTOP
                cell.border = BORDER_ALL
            # Code colour stripe
            did = (rec.get("Defect ID") or "").strip()
            be  = (rec.get("Building element") or "").strip()
            cidx = headers.index("Defect ID") + 1 if "Defect ID" in headers else None
            beidx= headers.index("Building element") + 1 if "Building element" in headers else None
            if cidx:
                code_cell = s_def.cell(i, cidx)
                colour = REGIME_COLOURS.get(be, None)
                if not colour and did:
                    code_letter = re.match(r"^[A-Z]+", did)
                    if code_letter:
                        colour = {
                            "W": "1E90FF", "F": "B22222", "S": "8B5A3C",
                            "BE": "DAA520", "ES": "20B2AA",
                        }.get(code_letter.group(0), None)
                if colour:
                    code_cell.fill = hdr_fill(colour)
                    code_cell.font = font(WHITE, bold=True)
                    code_cell.alignment = CENTER
            if beidx:
                be_cell = s_def.cell(i, beidx)
                colour = REGIME_COLOURS.get(be, None)
                if colour:
                    be_cell.fill = hdr_fill(colour); be_cell.font = font(WHITE, bold=True)
                    be_cell.alignment = CENTER
        # Width caps
        autosize(s_def, max_widths={
            "A": 14, "B": 26, "C": 28, "D": 60, "E": 60, "F": 50,
            "G": 50, "H": 50, "I": 50, "J": 18, "K": 50, "L": 35, "M": 16, "N": 30
        })
        add_filter_and_freeze(s_def)

    # =====================================================================
    # SHEET 08 — Quick Lookup by Element (pivoted defect+crosswalk)
    # =====================================================================
    s_qk = wb.create_sheet("07 Quick Lookup by Element")
    headers = ["Building Element / Regime", "Sub-Category", "Defect Codes",
               "Indicative NCC 2022 Refs", "Indicative AS Refs", "Defect Count"]
    write_header_row(s_qk, 1, headers, "1E90FF", WHITE)

    by_element = {}
    for d in defects:
        be = (d.get("Building element") or "Unspecified").strip()
        sc = (d.get("Sub-categories") or "").strip()
        key = (be, sc)
        e = by_element.setdefault(key, {"codes": [], "ncc22": set(), "as_refs": set()})
        e["codes"].append((d.get("Defect ID") or "").strip())
        ncc22 = d.get("NCC 2022 Reference") or ""
        if ncc22:
            for token in re.findall(r"\b(?:Volume\s*\w+|Section\s*\w+|Part\s*\w+|Specification\s*\w+|Clause\s*\w+|\d+\.\d+(?:\.\d+)*|[A-Z]+\d+[A-Z]?\d*)\b", ncc22):
                e["ncc22"].add(token)
        # mine AS codes
        for src in (d.get("NCC 2019 Reference"), d.get("NCC 2022 Reference"),
                    d.get("Requirement for standard of work"),
                    d.get("Reason why it is a serious defect")):
            if src and isinstance(src, str):
                for as_code in re.findall(r"AS(?:/NZS)?\s*\d+(?:\.\d+)*(?::\d{4})?", src):
                    e["as_refs"].add(as_code.strip())

    rownum = 2
    for (be, sc), v in sorted(by_element.items(), key=lambda kv: (kv[0][0], kv[0][1])):
        bcell = s_qk.cell(rownum, 1, be)
        colour = REGIME_COLOURS.get(be, None)
        if colour:
            bcell.fill = hdr_fill(colour); bcell.font = font(WHITE, bold=True); bcell.alignment = CENTER
        else:
            bcell.font = font(NAVY, bold=True)
        s_qk.cell(rownum, 2, sc)
        s_qk.cell(rownum, 3, ", ".join(sorted(set([c for c in v["codes"] if c])))).alignment = LEFTTOP
        s_qk.cell(rownum, 4, ", ".join(sorted(v["ncc22"]))).alignment = LEFTTOP
        s_qk.cell(rownum, 5, ", ".join(sorted(v["as_refs"]))).alignment = LEFTTOP
        s_qk.cell(rownum, 6, len([c for c in v["codes"] if c])).alignment = CENTER
        for col in range(1, len(headers)+1):
            s_qk.cell(rownum, col).border = BORDER_ALL
        rownum += 1
    autosize(s_qk, max_widths={"A": 32, "B": 28, "C": 50, "D": 50, "E": 40, "F": 14})
    add_filter_and_freeze(s_qk)

    # =====================================================================
    # SHEET 09 — Reference Standards Index (alphabetical by code)
    # =====================================================================
    s_ix = wb.create_sheet("08 Reference Standards Index")
    headers = ["Standard Code", "Title", "NCC Version(s)", "Year(s)", "Cross-references"]
    write_header_row(s_ix, 1, headers, GOLD, WHITE)
    by_code = {}
    for r in refstd:
        key = r["code"].strip()
        if not key: continue
        d = by_code.setdefault(key, {"titles": set(), "vers": set(), "years": set()})
        if r["title"]: d["titles"].add(r["title"])
        if r["ncc_version"]: d["vers"].add(r["ncc_version"])
        if r["year"]: d["years"].add(r["year"])
    # add AS in use codes
    for r in asuse:
        key = r["code"].strip()
        if not key: continue
        d = by_code.setdefault(key, {"titles": set(), "vers": set(), "years": set()})
        if r["title"]: d["titles"].add(r["title"])
        d["vers"].add("AS in use list")
    for i, (code, d) in enumerate(sorted(by_code.items()), 2):
        s_ix.cell(i, 1, code).font = font(NAVY, bold=True)
        s_ix.cell(i, 2, "; ".join(sorted(d["titles"]))).alignment = LEFTTOP
        vers_str = ", ".join(sorted(d["vers"]))
        vc = s_ix.cell(i, 3, vers_str); vc.alignment = CENTER
        if "NCC2022" in vers_str and "NCC2019" in vers_str:
            vc.fill = hdr_fill("C6EFCE")
        elif "NCC2022" in vers_str:
            vc.fill = hdr_fill("EFE3CC")
        elif "NCC2019" in vers_str:
            vc.fill = hdr_fill("FFE699")
        s_ix.cell(i, 4, ", ".join(sorted(d["years"]))).alignment = CENTER
        s_ix.cell(i, 5, "")  # reserved for future cross-ref enrichment
        for col in range(1, len(headers)+1):
            s_ix.cell(i, col).border = BORDER_ALL
    autosize(s_ix, max_widths={"A": 30, "B": 80, "C": 22, "D": 16, "E": 40})
    add_filter_and_freeze(s_ix)

    # =====================================================================
    # SHEET 10 — Section Colour Legend
    # =====================================================================
    s_lg = wb.create_sheet("09 Section Colour Legend")
    s_lg.sheet_view.showGridLines = False
    s_lg.column_dimensions['A'].width = 3
    s_lg.column_dimensions['B'].width = 18
    s_lg.column_dimensions['C'].width = 36
    s_lg.column_dimensions['D'].width = 16

    s_lg.merge_cells("B2:D2"); h = s_lg["B2"]
    h.value = "VISUAL LEGEND — NCC Sections, Clause Types, Regimes"
    h.font = font(NAVY, bold=True, size=16); h.alignment = LEFT

    legend_blocks = [
        ("NCC VOLUME ONE — SECTIONS", [
            ("A — General provisions",         "5B6C8C"),
            ("B — Structure",                  "8B5A3C"),
            ("C — Fire resistance",            "B22222"),
            ("D — Access & egress",            "FF8C00"),
            ("E — Services & equipment",       "DAA520"),
            ("F — Health & amenity",           "1E90FF"),
            ("G — Ancillary provisions",       "228B22"),
            ("I — Special use buildings",      "9370DB"),
            ("J — Energy efficiency",          "20B2AA"),
            ("S — Specifications",             "696969"),
        ]),
        ("NCC VOLUME TWO — HOUSING PROVISIONS", [
            ("Vol 2 (HP) Sections 2-13",       "4F6228"),
        ]),
        ("CLAUSE TYPE", [
            ("Objective (O)",                  "B0C4DE"),
            ("Functional Statement (F)",       "C2E0FF"),
            ("Performance Requirement (P)",    "F4B084"),
            ("Verification Method (V)",        "FFD966"),
            ("Deemed-to-Satisfy (D)",          "C6EFCE"),
            ("Specification Clause (Sx C)",    "D9D9D9"),
            ("Governing / General (G)",        "E7E6E6"),
        ]),
        ("DEFECT REGIMES (NSW Building Commissioner)", [
            ("W — Waterproofing",              "1E90FF"),
            ("F — Fire safety",                "B22222"),
            ("S — Structure",                  "8B5A3C"),
            ("BE — Building Enclosure",        "DAA520"),
            ("ES — Essential Services",        "20B2AA"),
        ]),
        ("NCC VERSION HIGHLIGHT", [
            ("Standard referenced by NCC 2022", "C6EFCE"),
            ("Standard referenced by NCC 2019", "FFE699"),
            ("Referenced in BOTH editions",     "EFE3CC"),
        ]),
    ]
    r = 4
    for title, items in legend_blocks:
        s_lg.merge_cells(f"B{r}:D{r}")
        c = s_lg.cell(r, 2, title); c.fill = hdr_fill(NAVY)
        c.font = font(WHITE, bold=True, size=11); c.alignment = LEFT
        r += 1
        for label, hex_ in items:
            s_lg.cell(r, 2, label).font = font(NAVY)
            sw = s_lg.cell(r, 3, ""); sw.fill = hdr_fill(hex_)
            s_lg.cell(r, 4, "#" + hex_).font = font(GREY_HDR, italic=True, size=9)
            s_lg.cell(r, 4).alignment = CENTER
            r += 1
        r += 1

    # =====================================================================
    # Cover hyperlinks (now that all sheets exist)
    # =====================================================================
    cover = wb["Cover"]
    sheet_link_map = [
        (7,  "01 NCC 2022 Vol 1 Clauses"),
        (8,  "02 NCC 2022 Vol 2 (HP)"),
        (9,  "03 AS Standards Master"),
        (10, "04 AS in Use (categorised)"),
        (11, "05 Topic-NCC-AS Crosswalk"),
        (12, "06 Master Defect Library"),
        (13, "07 Quick Lookup by Element"),
        (14, "08 Reference Standards Index"),
        (15, "09 Section Colour Legend"),
    ]
    for row_num, sheet_name in sheet_link_map:
        cell = cover.cell(row_num, 2)
        cell.hyperlink = f"#'{sheet_name}'!A1"
        cell.font = font("0563C1", bold=True, size=11)
    print(f"[ ] Saving …")
    wb.save(OUT)
    size_kb = OUT.stat().st_size / 1024
    print(f"\n[OK] Saved {OUT}  ({size_kb:,.0f} KB)")

if __name__ == "__main__":
    main()
